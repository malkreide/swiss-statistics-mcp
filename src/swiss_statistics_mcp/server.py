"""
Swiss Statistics MCP Server
============================
Provides access to Swiss Federal Statistical Office (BFS) data
via the STAT-TAB PxWeb API (pxweb.bfs.admin.ch).

All 682 statistical datasets across 21 themes are accessible:
Bevölkerung, Bildung, Arbeit, Gesundheit, Politik, and more.

No authentication required. Open data under BFS usage terms.
"""

from __future__ import annotations

import asyncio
import csv
import functools
import html
import io
import json
import logging
import os
import re
import sys
import time
import uuid
from collections.abc import Callable
from datetime import date
from typing import Any, Literal
from urllib.parse import quote_plus, urlencode

import httpx
from mcp.server.fastmcp import FastMCP
from pydantic import BaseModel, ConfigDict, Field
from tenacity import (
    AsyncRetrying,
    RetryError,
    retry_if_exception,
    stop_after_attempt,
    wait_exponential,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

BFS_API_BASE = "https://www.pxweb.bfs.admin.ch/api/v1"
DEFAULT_LANGUAGE = "de"
HTTP_TIMEOUT = 30.0
CATALOG_CACHE_TTL = 3600  # 1 hour
METADATA_CACHE_TTL = 3600  # 1 hour
FANOUT_CONCURRENCY = 5  # parallel metadata fetches per fan-out tool call

# Defense-in-depth allowlist for BFS table identifiers. The value is
# interpolated into URL path segments (`{BFS_API_BASE}/{lang}/{dbid}/...`),
# so we reject anything that's not the BFS-canonical
# `px-x-1234567890_123` shape before the request leaves the process.
# httpx encodes path segments correctly even without this, but having the
# guard at the input boundary (a) shortens error feedback, and (b) keeps
# any future cache key or log statement from carrying unexpected
# characters. (SEC-008)
BFS_TABLE_ID_PATTERN = r"^px-[a-z]-\d{8,12}_\d{1,4}$"

# Retry policy for outbound BFS calls. Defaults absorb the typical 503/504
# blips from PxWeb without inflating tail latency on the common-case happy
# path. Override via env for tests or aggressive deployments.
RETRY_MAX_ATTEMPTS = int(os.environ.get("MCP_RETRY_MAX_ATTEMPTS", "3"))
RETRY_WAIT_INITIAL = float(os.environ.get("MCP_RETRY_WAIT_INITIAL", "0.5"))
RETRY_WAIT_MAX = float(os.environ.get("MCP_RETRY_WAIT_MAX", "4.0"))

# ---------------------------------------------------------------------------
# Reference layer: Amtliches Gemeindeverzeichnis (BFS AGVCH) + HSSO
# ---------------------------------------------------------------------------
#
# The commune register is the portfolio's REFERENCE LAYER: the BFS commune
# number is the join key that ties statistics, geo, education and health
# data together. The AGVCH REST service (rest_api_de.pdf, v2026.06.x) is a
# clean, versioned, No-Auth API — snapshot / correspondances / mutations —
# so these tools use Architecture A (live-API-only) with a TTL cache.
#
# Historical Statistics of Switzerland (hsso.ch) offers no API, only static
# per-table XLSX dumps at stable URLs → Architecture C (dump-only). Its
# licence is CC BY-NC-SA 3.0 (NonCommercial), which differs from the OGD
# baseline of the rest of this server, so every HSSO response carries an
# explicit NonCommercial notice.

AGVCH_API_BASE = "https://www.agvchapp.bfs.admin.ch/api/communes"
HSSO_BASE = "https://hsso.ch"
LINDAS_MUNICIPALITY_URI = "https://ld.admin.ch/municipality/{bfs}"

AGVCH_ATTRIBUTION = (
    "Amtliches Gemeindeverzeichnis der Schweiz (BFS) — "
    "https://www.bfs.admin.ch/bfs/de/home/grundlagen/agvch.html. "
    "Open Government Data, freie Weiterverwendung."
)
HSSO_ATTRIBUTION = (
    "Historische Statistik der Schweiz (HSSO) — https://hsso.ch. "
    "Lizenz CC BY-NC-SA 3.0: Namensnennung erforderlich, keine kommerzielle Nutzung."
)
HSSO_NONCOMMERCIAL_NOTE = (
    "⚠️ HSSO-Daten stehen unter CC BY-NC-SA 3.0 (NonCommercial). "
    "Nur für nicht-kommerzielle Nutzung mit Quellenangabe."
)

# Commune snapshots change only at year boundaries and mutations are
# historical (append-only), so a day-long cache is safe and collapses the
# repeated lookup/list flows to a single fetch per (date) key.
COMMUNE_CACHE_TTL = int(os.environ.get("MCP_COMMUNE_CACHE_TTL", "86400"))  # 24h
HSSO_INDEX_TTL = int(os.environ.get("MCP_HSSO_INDEX_TTL", "86400"))  # 24h

# Snapshot Level codes → human labels. (Finding: the live CSV header uses
# Inscription/Radiation/Rec_Type_fr, NOT the Einschreibung/Streichung names
# printed in rest_api_de.pdf — parse against the live header, not the doc.)
_AGVCH_LEVEL_LABELS = {"1": "Kanton", "2": "Bezirk", "3": "Gemeinde"}


# ---------------------------------------------------------------------------
# Logging (OBS-001, OBS-002, OBS-003, SEC-014)
# ---------------------------------------------------------------------------
#
# Logs are emitted to stderr as JSON so cloud log aggregators (Render, etc.)
# can index per-tool latency and errors. stdout is reserved for the MCP
# protocol on stdio transport — never write logs there.

class _JsonFormatter(logging.Formatter):
    """JSON formatter that accepts dict-shaped records and falls back to
    plain messages for library logs."""

    def format(self, record: logging.LogRecord) -> str:
        payload: dict[str, Any] = {
            "ts": self.formatTime(record, "%Y-%m-%dT%H:%M:%S"),
            "level": record.levelname,
            "logger": record.name,
        }
        if isinstance(record.msg, dict):
            payload.update(record.msg)
        else:
            payload["msg"] = record.getMessage()
        if record.exc_info:
            payload["exc_info"] = self.formatException(record.exc_info)
        return json.dumps(payload, ensure_ascii=False)


def _configure_logger() -> logging.Logger:
    log = logging.getLogger("swiss_statistics_mcp")
    if not log.handlers:
        handler = logging.StreamHandler(sys.stderr)
        handler.setFormatter(_JsonFormatter())
        log.addHandler(handler)
        log.propagate = False
    log.setLevel(os.environ.get("MCP_LOG_LEVEL", "INFO").upper())
    return log


_LOGGER = _configure_logger()


def _logged_tool(tool_name: str) -> Callable[[Callable[..., Any]], Callable[..., Any]]:
    """Wrap an MCP tool with start/end JSON logs.

    Emits one `tool_start` and one `tool_end` event per call, both tagged
    with a per-call `rid` (correlation id) and `duration_ms` on the end
    event. Param **keys** (not values) are logged to avoid surfacing user
    input — values may contain locale strings or query text, and keeping
    the field list keys-only sidesteps any future privacy concern when
    the same plumbing is reused on PII data.
    """
    def decorator(fn: Callable[..., Any]) -> Callable[..., Any]:
        @functools.wraps(fn)
        async def wrapper(params: BaseModel, *args: Any, **kwargs: Any) -> Any:
            rid = uuid.uuid4().hex[:8]
            t0 = time.monotonic()
            try:
                param_keys = sorted(params.model_dump(exclude_none=True).keys())
            except Exception:
                param_keys = []
            _LOGGER.info({
                "event": "tool_start",
                "tool": tool_name,
                "rid": rid,
                "params_keys": param_keys,
            })
            try:
                result = await fn(params, *args, **kwargs)
                _LOGGER.info({
                    "event": "tool_end",
                    "tool": tool_name,
                    "rid": rid,
                    "status": "ok",
                    "duration_ms": int((time.monotonic() - t0) * 1000),
                })
                return result
            except Exception as e:
                _LOGGER.info({
                    "event": "tool_end",
                    "tool": tool_name,
                    "rid": rid,
                    "status": "error",
                    "error_type": type(e).__name__,
                    "duration_ms": int((time.monotonic() - t0) * 1000),
                })
                raise
        return wrapper
    return decorator

BFS_THEMES: dict[str, str] = {
    "01": "Bevölkerung",
    "02": "Raum und Umwelt",
    "03": "Arbeit und Erwerb",
    "04": "Volkswirtschaft",
    "05": "Preise",
    "06": "Industrie und Dienstleistungen",
    "07": "Land- und Forstwirtschaft",
    "08": "Energie",
    "09": "Bau- und Wohnungswesen",
    "10": "Tourismus",
    "11": "Mobilität und Verkehr",
    "12": "Geld, Banken, Versicherungen",
    "13": "Soziale Sicherheit",
    "14": "Gesundheit",
    "15": "Bildung und Wissenschaft",
    "16": "Kultur, Medien, Informationsgesellschaft",
    "17": "Politik",
    "18": "Öffentliche Verwaltung und Finanzen",
    "19": "Kriminalität und Strafrecht",
    "20": "Wirtschaftliche und soziale Situation",
    "21": "Nachhaltige Entwicklung",
}

# Curated high-value tables for Schulamt / education context
FEATURED_TABLES: dict[str, str] = {
    "px-x-1504000000_173": "Lehrkräfte nach Schuljahr, Kanton und Bildungsstufe",
    "px-x-1504000000_172": "Lehrkräfte nach Schuljahr, Kanton und Staatsangehörigkeit",
    "px-x-0102010000_101": "Ständige Wohnbevölkerung nach Kanton, Alter und Geschlecht",
    "px-x-1509090000_101": "Szenarien Sekundarstufe II: Entwicklung Schülerzahlen",
    "px-x-1509090000_113": "Szenarien Hochschulen: Entwicklung Studierendenzahlen",
    "px-x-1502020100_101": "Schülerinnen und Schüler nach Bildungsstufe und Kanton",
    "px-x-1503040100_101": "Abschlüsse Sekundarstufe II nach Kanton",
    "px-x-1506020000_114": "Stipendien und Darlehen nach Kanton",
    "px-x-1703030000_101": "Nationalratswahlen: Resultate nach Kanton",
    "px-x-1703030000_100": "Volksabstimmungen: Resultate",
    "px-x-0301000000_101": "Erwerbstätige nach Wirtschaftszweig",
    "px-x-1302020000_101": "Sozialhilfe: Quoten nach Kanton",
}

# ---------------------------------------------------------------------------
# In-memory catalog cache
# ---------------------------------------------------------------------------

_catalog_cache: dict[str, Any] = {}
_catalog_timestamp: float = 0.0


# ---------------------------------------------------------------------------
# HTTP helpers (SEC-018, SCALE-002)
# ---------------------------------------------------------------------------
#
# All outbound BFS calls run through `_get`/`_post` and inherit the same
# retry policy: up to 3 attempts on transient errors (5xx, 429, network
# errors), with exponential backoff. 4xx errors are surfaced immediately
# because retrying client errors only wastes upstream quota.

_TRANSIENT_HTTP_STATUSES = {429, 500, 502, 503, 504}


def _is_transient(exc: BaseException) -> bool:
    if isinstance(
        exc,
        (httpx.TimeoutException, httpx.ConnectError, httpx.ReadError, httpx.WriteError),
    ):
        return True
    if isinstance(exc, httpx.HTTPStatusError):
        return exc.response.status_code in _TRANSIENT_HTTP_STATUSES
    return False


async def _retrying_http(coro_factory: Callable[[], Any]) -> Any:
    """Run `coro_factory()` with retry on transient errors.

    `coro_factory` is a zero-arg callable returning a coroutine; we call it
    fresh on each retry so a new `AsyncClient` is opened per attempt
    (httpx clients are not safe to reuse after errors).
    """
    try:
        async for attempt in AsyncRetrying(
            stop=stop_after_attempt(RETRY_MAX_ATTEMPTS),
            wait=wait_exponential(
                multiplier=RETRY_WAIT_INITIAL,
                min=RETRY_WAIT_INITIAL,
                max=RETRY_WAIT_MAX,
            ),
            retry=retry_if_exception(_is_transient),
            reraise=True,
        ):
            with attempt:
                return await coro_factory()
    except RetryError as e:  # pragma: no cover — reraise=True usually raises the wrapped exc
        raise e.last_attempt.exception() from e


async def _get(url: str) -> Any:
    """Perform a GET request and return parsed JSON, with retry on transient errors."""
    async def _do() -> Any:
        async with httpx.AsyncClient(timeout=HTTP_TIMEOUT) as client:
            resp = await client.get(url)
            resp.raise_for_status()
            return resp.json()
    return await _retrying_http(_do)


async def _post(url: str, body: dict[str, Any]) -> Any:
    """Perform a POST request and return parsed JSON, with retry on transient errors."""
    async def _do() -> Any:
        async with httpx.AsyncClient(timeout=HTTP_TIMEOUT) as client:
            resp = await client.post(url, json=body)
            resp.raise_for_status()
            return resp.json()
    return await _retrying_http(_do)


async def _get_text(url: str) -> str:
    """Perform a GET request and return the raw response text, with retry.

    Used for the AGVCH CSV endpoints and HSSO HTML pages, which are not JSON.
    Shares the same transient-error retry policy as `_get`/`_post`.
    """
    async def _do() -> str:
        async with httpx.AsyncClient(timeout=HTTP_TIMEOUT, follow_redirects=True) as client:
            resp = await client.get(url)
            resp.raise_for_status()
            return resp.text
    return await _retrying_http(_do)


# ---------------------------------------------------------------------------
# AGVCH commune register helpers (Architecture A: live REST + TTL cache)
# ---------------------------------------------------------------------------
#
# The AGVCH REST service speaks DD-MM-YYYY; the tools accept portfolio-native
# ISO dates (YYYY-MM-DD) and convert at the boundary so every server in the
# portfolio exposes the same date contract.

_snapshot_cache: dict[str, tuple[float, list[dict[str, str]]]] = {}


def _iso_to_agvch(iso_date: str) -> str:
    """'2025-01-01' → '01-01-2025' (AGVCH's DD-MM-YYYY)."""
    y, m, d = iso_date.split("-")
    return f"{d}-{m}-{y}"


async def _fetch_agvch_csv(endpoint: str, params: dict[str, str]) -> list[dict[str, str]]:
    """Fetch an AGVCH CSV endpoint and parse it against its live header."""
    url = f"{AGVCH_API_BASE}/{endpoint}?{urlencode(params)}"
    text = await _get_text(url)
    return list(csv.DictReader(io.StringIO(text)))


async def _fetch_snapshot(agvch_date: str) -> tuple[list[dict[str, str]], bool]:
    """Return (rows, from_cache) for the commune snapshot at a DD-MM-YYYY date.

    A snapshot lists cantons (Level 1), districts (Level 2) and communes
    (Level 3) with their `Parent` links, so canton membership is derivable
    from the snapshot alone — no need for the 2.5 MB `levels` endpoint.
    """
    now = time.time()
    hit = _snapshot_cache.get(agvch_date)
    if hit is not None and (now - hit[0]) < COMMUNE_CACHE_TTL:
        return hit[1], True
    rows = await _fetch_agvch_csv("snapshot", {"date": agvch_date, "format": "csv"})
    _snapshot_cache[agvch_date] = (now, rows)
    return rows, False


def _index_by_hist(rows: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    """Index snapshot rows by HistoricalCode as a MULTIMAP.

    Finding: HistoricalCode is NOT globally unique in the snapshot — e.g.
    code 10078 is both ZH's 'Bezirk Horgen' (Level 2) and VS's commune
    'Vionnaz' (Level 3). A plain dict would silently collapse the two and
    resolve Parent links to the wrong entity, so the Parent link must be
    disambiguated by tier (see `_parent_row`).
    """
    idx: dict[str, list[dict[str, str]]] = {}
    for r in rows:
        h = r.get("HistoricalCode")
        if h:
            idx.setdefault(h, []).append(r)
    return idx


def _level_int(row: dict[str, str]) -> int:
    try:
        return int(row.get("Level") or 0)
    except ValueError:
        return 0


def _parent_row(
    row: dict[str, str], by_hist: dict[str, list[dict[str, str]]]
) -> dict[str, str] | None:
    """Resolve a row's Parent to the candidate exactly one tier up.

    Parent references a HistoricalCode, which may be shared across levels;
    the true parent is the candidate with the highest Level that is still
    strictly above (lower Level number than) the current row.
    """
    parent = row.get("Parent") or ""
    if not parent:
        return None
    cur_level = _level_int(row)
    best: dict[str, str] | None = None
    best_level = -1
    for cand in by_hist.get(parent, []):
        lv = _level_int(cand)
        if 0 < lv < cur_level and lv > best_level:
            best, best_level = cand, lv
    return best


def _climb_to_canton(
    row: dict[str, str], by_hist: dict[str, list[dict[str, str]]]
) -> dict[str, str] | None:
    """Walk the Parent chain up to the Level-1 (canton) row, or None.

    commune → district → canton (some cantons skip the district tier).
    Guards against cycles / dangling parents.
    """
    cur: dict[str, str] | None = row
    seen: set[tuple[str, str]] = set()
    while cur is not None and cur.get("Level") != "1":
        key = (cur.get("HistoricalCode", ""), cur.get("Level", ""))
        if key in seen:
            return None
        seen.add(key)
        cur = _parent_row(cur, by_hist)
    return cur


def _commune_entry(
    row: dict[str, str], by_hist: dict[str, list[dict[str, str]]]
) -> CommuneEntry:
    """Build a CommuneEntry from a snapshot row, enriched with canton + URI."""
    level = row.get("Level", "")
    canton_row = _climb_to_canton(row, by_hist)
    try:
        bfs = int(row.get("BfsCode", "") or 0)
    except ValueError:
        bfs = 0
    lindas = LINDAS_MUNICIPALITY_URI.format(bfs=bfs) if level == "3" else None
    return CommuneEntry(
        bfs_number=bfs,
        name=row.get("Name", ""),
        short_name=(row.get("ShortName") or None),
        historical_code=(row.get("HistoricalCode") or None),
        level=_AGVCH_LEVEL_LABELS.get(level, level or "?"),
        canton=(canton_row.get("Name") if canton_row else None),
        canton_abbr=(canton_row.get("ShortName") if canton_row else None),
        valid_from=(row.get("ValidFrom") or None),
        valid_to=(row.get("ValidTo") or None),  # empty ⇒ still valid
        lindas_uri=lindas,
    )


def _mutation_step(m: dict[str, str]) -> MutationStep:
    def _int_or_none(x: str | None) -> int | None:
        try:
            return int(x) if x else None
        except ValueError:
            return None

    return MutationStep(
        mutation_number=(m.get("MutationNumber") or None),
        mutation_date=(m.get("MutationDate") or None),
        initial_bfs=_int_or_none(m.get("InitialCode")),
        initial_name=(m.get("InitialName") or None),
        terminal_bfs=_int_or_none(m.get("TerminalCode")),
        terminal_name=(m.get("TerminalName") or None),
    )


# ---------------------------------------------------------------------------
# HSSO index helpers (Architecture C: dump-only, static XLSX at stable URLs)
# ---------------------------------------------------------------------------
#
# HSSO exposes no API. Each chapter index page lists its tables as
#   <a class="explorer-item" href="/de/2012/a/1a">
#     <div class="explorer-item__title">A.1a</div>
#     <div class="explorer-item__description">Areale nach Kantonen</div></a>
# and each table has a stable XLSX at /get/{CHAPTER}.{NN}{suffix}.xlsx
# (numeric part zero-padded to two digits). We build a searchable title
# index by scraping the 20 chapter pages once and caching it.

_hsso_index_cache: dict[str, tuple[float, list[HistoricalSeriesEntry]]] = {}

_HSSO_CHAPTERS = "abcdefghijklmnopqrst"

_HSSO_ITEM_RE = re.compile(
    r'href="(/de/2012/([a-z])/([a-z0-9]+))"[^>]*>\s*'
    r'<div class="explorer-item__title">([^<]+)</div>\s*'
    r'<div class="explorer-item__description">([^<]*)</div>',
    re.IGNORECASE,
)


def _hsso_xlsx_path(chapter: str, table_id: str) -> str:
    """'a', '1a' → '/get/A.01a.xlsx' (numeric part zero-padded to 2 digits)."""
    m = re.match(r"(\d+)([a-z]*)$", table_id)
    if not m:
        return f"/get/{chapter.upper()}.{table_id}.xlsx"
    num, suffix = m.group(1), m.group(2)
    return f"/get/{chapter.upper()}.{int(num):02d}{suffix}.xlsx"


def _parse_hsso_chapter(html_text: str) -> list[HistoricalSeriesEntry]:
    entries: list[HistoricalSeriesEntry] = []
    for m in _HSSO_ITEM_RE.finditer(html_text):
        path, chapter, table_id, code, desc = m.groups()
        entries.append(
            HistoricalSeriesEntry(
                code=html.unescape(code).strip(),
                title=html.unescape(desc).strip(),
                chapter=chapter,
                page_url=f"{HSSO_BASE}{path}",
                xlsx_url=f"{HSSO_BASE}{_hsso_xlsx_path(chapter, table_id)}",
            )
        )
    return entries


async def _ensure_hsso_index() -> tuple[list[HistoricalSeriesEntry], bool]:
    """Return (index, from_cache). Scrapes the 20 chapter pages concurrently.

    A partial scrape (some chapters unreachable) is still returned but NOT
    cached, so a transient outage can't poison the index for 24h.
    """
    now = time.time()
    hit = _hsso_index_cache.get("index")
    if hit is not None and (now - hit[0]) < HSSO_INDEX_TTL:
        return hit[1], True

    sem = asyncio.Semaphore(FANOUT_CONCURRENCY)

    async def fetch_chapter(ch: str) -> list[HistoricalSeriesEntry]:
        async with sem:
            try:
                text = await _get_text(f"{HSSO_BASE}/de/2012/{ch}")
            except Exception:
                _LOGGER.warning("hsso chapter fetch failed for %s", ch, exc_info=True)
                return []
            return _parse_hsso_chapter(text)

    chapter_results = await asyncio.gather(
        *(fetch_chapter(c) for c in _HSSO_CHAPTERS)
    )
    complete = all(chapter_results)  # every chapter yielded at least one table
    index: list[HistoricalSeriesEntry] = [e for r in chapter_results for e in r]
    if index and complete:
        _hsso_index_cache["index"] = (now, index)
    return index, False


def _theme_code_from_dbid(dbid: str) -> str:
    """Extract the 2-digit BFS theme code from a database ID.

    E.g. 'px-x-1504000000_173' → '15' (Bildung und Wissenschaft)
    """
    # Format: px-x-{THEME}{rest}_{suffix}
    # The numeric part starts after 'px-x-'
    numeric_part = dbid.replace("px-x-", "")
    return numeric_part[:2]


def _format_table_url(dbid: str, lang: str = DEFAULT_LANGUAGE) -> str:
    return f"{BFS_API_BASE}/{lang}/{dbid}/{dbid}.px"


def _build_data_url(dbid: str, lang: str = DEFAULT_LANGUAGE) -> str:
    return f"{BFS_API_BASE}/{lang}/{dbid}/{dbid}.px"


# ---------------------------------------------------------------------------
# Metadata cache (SCALE-003)
# ---------------------------------------------------------------------------
#
# Table metadata (variables, value domains, last_updated) is effectively
# stable between BFS publishing windows. Caching by (dbid, lang) for an
# hour collapses the listing / detail / compare flows from N round-trips
# to 1 after the first warm-up.

_metadata_cache: dict[tuple[str, str], dict[str, Any]] = {}
_metadata_timestamps: dict[tuple[str, str], float] = {}


async def _fetch_metadata_cached(dbid: str, lang: str) -> dict[str, Any]:
    key = (dbid, lang)
    now = time.time()
    cached_at = _metadata_timestamps.get(key, 0.0)
    if key in _metadata_cache and (now - cached_at) < METADATA_CACHE_TTL:
        return _metadata_cache[key]
    meta = await _get(_format_table_url(dbid, lang))
    _metadata_cache[key] = meta
    _metadata_timestamps[key] = now
    return meta


# ---------------------------------------------------------------------------
# Catalog management
# ---------------------------------------------------------------------------

async def _ensure_catalog(lang: str = DEFAULT_LANGUAGE) -> dict[str, str]:
    """Return the full catalog {dbid: title}, with TTL-based caching.

    Since the PxWeb API doesn't support text search natively, we build
    a local index by fetching metadata for every table once per hour.
    This is done lazily when search_tables is called.
    """
    global _catalog_cache, _catalog_timestamp

    cache_key = f"catalog_{lang}"
    now = time.time()

    if cache_key in _catalog_cache and (now - _catalog_timestamp) < CATALOG_CACHE_TTL:
        return _catalog_cache[cache_key]

    # Fetch all database IDs
    url = f"{BFS_API_BASE}/{lang}/"
    all_dbs = await _get(url)

    catalog: dict[str, str] = {}
    # Fetch titles in batches to avoid overwhelming the API
    async with httpx.AsyncClient(timeout=HTTP_TIMEOUT) as client:
        for db in all_dbs:
            dbid = db["dbid"]
            try:
                meta_url = f"{BFS_API_BASE}/{lang}/{dbid}/{dbid}.px"
                resp = await client.get(meta_url)
                if resp.status_code == 200:
                    meta = resp.json()
                    catalog[dbid] = meta.get("title", dbid)
            except Exception:
                _LOGGER.warning(
                    "catalog metadata fetch failed for %s", dbid, exc_info=True
                )
                catalog[dbid] = dbid  # fallback to ID if metadata unavailable

    _catalog_cache[cache_key] = catalog
    _catalog_timestamp = now
    return catalog


# ---------------------------------------------------------------------------
# Pydantic input models
# ---------------------------------------------------------------------------

class ListThemesInput(BaseModel):
    model_config = ConfigDict(extra="forbid")
    lang: str = Field(
        default="de",
        description="Language code: 'de' (German), 'fr' (French), 'it' (Italian), 'en' (English)",
        pattern="^(de|fr|it|en)$",
    )


class ListTablesByThemeInput(BaseModel):
    model_config = ConfigDict(extra="forbid")
    theme_code: str = Field(
        ...,
        description=(
            "2-digit BFS theme code, e.g. '15' for Bildung, '01' for Bevölkerung. "
            "Use bfs_list_themes to see all codes."
        ),
        pattern="^\\d{2}$",
    )
    lang: str = Field(
        default="de",
        description="Language code: 'de', 'fr', 'it', 'en'",
        pattern="^(de|fr|it|en)$",
    )
    limit: int = Field(
        default=20,
        description="Maximum number of tables to return",
        ge=1,
        le=100,
    )


class SearchTablesInput(BaseModel):
    model_config = ConfigDict(extra="forbid")
    query: str = Field(
        ...,
        description=(
            "Keyword(s) to search in table titles. Examples: 'Lehrkräfte', "
            "'Bevölkerung Kanton', 'Schüler Bildungsstufe', 'Abstimmung'"
        ),
        min_length=2,
        max_length=100,
    )
    theme_code: str | None = Field(
        default=None,
        description="Optional: filter to a specific theme code, e.g. '15' for Bildung",
        pattern="^\\d{2}$",
    )
    lang: str = Field(
        default="de",
        description="Language code: 'de', 'fr', 'it', 'en'",
        pattern="^(de|fr|it|en)$",
    )
    limit: int = Field(
        default=10,
        description="Maximum number of results to return",
        ge=1,
        le=50,
    )


class GetTableMetadataInput(BaseModel):
    model_config = ConfigDict(extra="forbid")
    table_id: str = Field(
        ...,
        description=(
            "BFS table/database ID, e.g. 'px-x-1504000000_173'. "
            "Obtain from bfs_search_tables or bfs_list_tables_by_theme."
        ),
        pattern=BFS_TABLE_ID_PATTERN,
    )
    lang: str = Field(
        default="de",
        description="Language code: 'de', 'fr', 'it', 'en'",
        pattern="^(de|fr|it|en)$",
    )


class DimensionFilter(BaseModel):
    model_config = ConfigDict(extra="forbid")
    code: str = Field(..., description="Variable code from table metadata, e.g. 'Kanton'")
    values: list[str] = Field(
        ...,
        description=(
            "List of value codes to include. Get codes from bfs_get_table_metadata. "
            "Example: ['1', '2'] for Zürich and Bern"
        ),
        min_length=1,
    )


class GetDataInput(BaseModel):
    model_config = ConfigDict(extra="forbid")
    table_id: str = Field(
        ...,
        description="BFS table ID, e.g. 'px-x-1504000000_173'",
        pattern=BFS_TABLE_ID_PATTERN,
    )
    filters: list[DimensionFilter] | None = Field(
        default=None,
        description=(
            "Optional dimension filters to narrow results. "
            "Without filters, all combinations are returned (may be large). "
            "Each filter specifies a variable code and the values to include."
        ),
    )
    lang: str = Field(
        default="de",
        description="Language code: 'de', 'fr', 'it', 'en'",
        pattern="^(de|fr|it|en)$",
    )
    max_rows: int = Field(
        default=500,
        description="Maximum number of data rows to return (safety limit)",
        ge=1,
        le=5000,
    )


class GetEducationStatsInput(BaseModel):
    model_config = ConfigDict(extra="forbid")
    topic: str = Field(
        ...,
        description=(
            "Education topic to retrieve: "
            "'teachers' (Lehrkräfte nach Kanton), "
            "'students' (Schüler nach Bildungsstufe), "
            "'scenarios' (Szenarien Schülerzahlen Sek II), "
            "'scholarships' (Stipendien nach Kanton)"
        ),
        pattern="^(teachers|students|scenarios|scholarships)$",
    )
    canton: str | None = Field(
        default=None,
        description=(
            "Optional canton filter. Use canton name or index value from metadata. "
            "Examples: 'Zürich' → value '1', 'Bern / Berne' → '2'. "
            "Leave empty for all cantons."
        ),
    )
    lang: str = Field(
        default="de",
        description="Language code: 'de', 'fr', 'it', 'en'",
        pattern="^(de|fr|it|en)$",
    )


class GetPopulationInput(BaseModel):
    model_config = ConfigDict(extra="forbid")
    region: str = Field(
        default="Schweiz",
        description=(
            "Region to query. Options: 'Schweiz' (national), "
            "or a canton name like 'Zürich', 'Bern / Berne', 'Luzern'. "
            "Use the exact name from BFS metadata."
        ),
    )
    year: str | None = Field(
        default=None,
        description="Year to filter, e.g. '2024'. Leave empty for all available years.",
    )
    breakdown: str = Field(
        default="total",
        description=(
            "Breakdown type: 'total' (all ages combined), "
            "'age' (by single year of age), "
            "'gender' (by gender)"
        ),
        pattern="^(total|age|gender)$",
    )


class CompareCantonsInput(BaseModel):
    model_config = ConfigDict(extra="forbid")
    table_id: str = Field(
        ...,
        description="BFS table ID to compare across cantons",
        pattern=BFS_TABLE_ID_PATTERN,
    )
    canton_values: list[str] = Field(
        ...,
        description=(
            "List of canton value codes to compare. "
            "Get codes from bfs_get_table_metadata. "
            "Example: ['1', '2', '3'] for ZH, BE, LU. "
            "Use value '0' for Switzerland total."
        ),
        min_length=2,
        max_length=27,
    )
    additional_filters: list[DimensionFilter] | None = Field(
        default=None,
        description="Optional additional dimension filters beyond canton selection",
    )
    lang: str = Field(
        default="de",
        description="Language code: 'de', 'fr', 'it', 'en'",
        pattern="^(de|fr|it|en)$",
    )


# ---------------------------------------------------------------------------
# Output models (ARCH-005, ARCH-009)
# ---------------------------------------------------------------------------
#
# Every tool now returns a typed Pydantic model rather than a JSON string.
# FastMCP serialises these as structured content so clients can render
# fields directly and follow-up tool calls can be typed against the schema.
#
# Each result includes `error: str | None` and `hint: str | None` at the
# top level for explicit error discrimination — clients check
# `result.error is None` to know the call succeeded. On error the data
# fields are None.
#
# Data-returning tools (get_data, education_stats, population,
# compare_cantons) carry `truncated: bool`, `rows_total: int`, and
# `rows_returned: int` so clients no longer need to parse a German
# `warning` prose string to discover that rows were capped. (ARCH-009)


class DimensionInfo(BaseModel):
    id: str
    label: str
    n_values: int


class ThemeEntry(BaseModel):
    code: str
    name: str
    dataset_count: int
    filter_hint: str | None = None


class TableEntry(BaseModel):
    table_id: str
    title: str
    last_updated: str | None = None
    n_variables: int | None = None
    featured: bool | None = None


class SearchResultEntry(BaseModel):
    table_id: str
    title: str
    theme_code: str
    theme_name: str | None = None
    featured: bool | None = None


class VariableValue(BaseModel):
    code: str
    label: str


class VariableInfo(BaseModel):
    code: str
    label: str
    n_values: int
    values: list[VariableValue]
    more_values: int = 0  # how many values were truncated from `values` (display cap)


class FeaturedDatasetEntry(BaseModel):
    table_id: str
    title: str
    theme_code: str | None = None
    theme_name: str | None = None
    schulamt_relevanz: str | None = None


class ListThemesResult(BaseModel):
    error: str | None = None
    hint: str | None = None
    total_datasets: int | None = None
    themes: list[ThemeEntry] | None = None
    note: str | None = None


class ListTablesByThemeResult(BaseModel):
    error: str | None = None
    hint: str | None = None
    theme_code: str | None = None
    theme_name: str | None = None
    total_in_theme: int | None = None
    returned: int | None = None
    tables: list[TableEntry] | None = None
    next_step: str | None = None


class SearchTablesResult(BaseModel):
    error: str | None = None
    hint: str | None = None
    query: str | None = None
    total_matches: int | None = None
    results: list[SearchResultEntry] | None = None
    next_step: str | None = None


class TableMetadataResult(BaseModel):
    error: str | None = None
    hint: str | None = None
    table_id: str | None = None
    title: str | None = None
    source: str | None = None
    last_updated: str | None = None
    theme_code: str | None = None
    theme_name: str | None = None
    language: str | None = None
    n_variables: int | None = None
    variables: list[VariableInfo] | None = None
    usage_hint: str | None = None


class DataTableResult(BaseModel):
    error: str | None = None
    hint: str | None = None
    table_id: str | None = None
    title: str | None = None
    source: str | None = None
    updated: str | None = None
    language: str | None = None
    dimensions: list[DimensionInfo] | None = None
    rows_total: int | None = None
    rows_returned: int | None = None
    truncated: bool = False
    rows: list[dict[str, Any]] | None = None
    note: str | None = None
    # bfs_compare_cantons extras
    cantons_compared: list[str] | None = None
    canton_variable: str | None = None
    # bfs_education_stats / bfs_population extras
    topic: str | None = None
    topic_description: str | None = None
    canton: str | None = None
    canton_filter: str | None = None
    region: str | None = None
    breakdown: str | None = None
    year: str | None = None


class FeaturedDatasetsResult(BaseModel):
    error: str | None = None
    hint: str | None = None
    total: int | None = None
    featured_datasets: list[FeaturedDatasetEntry] | None = None
    quick_start: str | None = None


# ---------------------------------------------------------------------------
# Reference layer models: communes (AGVCH) + historical series (HSSO)
# ---------------------------------------------------------------------------
#
# Every reference-layer response carries `source` (attribution) and
# `provenance` (live_api | cached) so downstream consumers never lose the
# data lineage — the README is not passed along, the envelope is.


class CommuneEntry(BaseModel):
    bfs_number: int
    name: str
    short_name: str | None = None
    historical_code: str | None = None
    level: str  # Gemeinde | Bezirk | Kanton
    canton: str | None = None
    canton_abbr: str | None = None
    valid_from: str | None = None
    valid_to: str | None = None  # None ⇒ still valid at the queried date
    lindas_uri: str | None = None  # stable LINDAS/Linked-Data identifier


class SuccessorEntry(BaseModel):
    bfs_number: int
    name: str
    lindas_uri: str | None = None


class MutationStep(BaseModel):
    mutation_number: str | None = None
    mutation_date: str | None = None
    initial_bfs: int | None = None
    initial_name: str | None = None
    terminal_bfs: int | None = None
    terminal_name: str | None = None


class HistoricalSeriesEntry(BaseModel):
    code: str  # HSSO table code, e.g. "A.1a"
    title: str
    chapter: str
    page_url: str
    xlsx_url: str  # stable static download


class LookupCommuneInput(BaseModel):
    model_config = ConfigDict(extra="forbid")
    name_or_bfs_number: str = Field(
        ...,
        description=(
            "Commune name (or substring, e.g. 'Wädenswil') or BFS number "
            "(e.g. '293'). Numeric input is matched exactly against the "
            "BFS number; text is matched case-insensitively against names."
        ),
        min_length=1,
        max_length=100,
    )
    valid_at_date: str = Field(
        default_factory=lambda: date.today().isoformat(),
        description=(
            "ISO date (YYYY-MM-DD): return the commune as it existed on this "
            "date. Default: today. Use a historical date to see former communes."
        ),
        pattern=r"^\d{4}-\d{2}-\d{2}$",
    )


class ResolveHistoricalCommuneInput(BaseModel):
    model_config = ConfigDict(extra="forbid")
    bfs_number: int = Field(
        ...,
        description=(
            "Historical BFS commune number to resolve, e.g. 133 (old Horgen) "
            "or 132 (Hirzel, dissolved). The number the old statistics use."
        ),
        ge=1,
        le=9999,
    )
    from_date: str = Field(
        ...,
        description=(
            "ISO date (YYYY-MM-DD) the old data belongs to — the commune must "
            "have existed on this date."
        ),
        pattern=r"^\d{4}-\d{2}-\d{2}$",
    )
    to_date: str = Field(
        default_factory=lambda: date.today().isoformat(),
        description="Target ISO date (YYYY-MM-DD) to map onto. Default: today.",
        pattern=r"^\d{4}-\d{2}-\d{2}$",
    )


class ListCommunesInput(BaseModel):
    model_config = ConfigDict(extra="forbid")
    canton: str = Field(
        ...,
        description=(
            "Canton abbreviation (e.g. 'ZH', 'BE') or name (e.g. 'Zürich'). "
            "The canton join key for the whole portfolio."
        ),
        min_length=2,
        max_length=40,
    )
    valid_at_date: str = Field(
        default_factory=lambda: date.today().isoformat(),
        description="ISO date (YYYY-MM-DD). Default: today.",
        pattern=r"^\d{4}-\d{2}-\d{2}$",
    )


class SearchHistoricalSeriesInput(BaseModel):
    model_config = ConfigDict(extra="forbid")
    topic: str = Field(
        ...,
        description=(
            "Keyword(s) to search HSSO table titles, e.g. 'Bevölkerung', "
            "'Preise', 'Verkehr', 'Landwirtschaft'. All terms must match."
        ),
        min_length=2,
        max_length=100,
    )
    period: str | None = Field(
        default=None,
        description=(
            "Optional period hint, e.g. '1850-1900'. Informational only — HSSO "
            "does not expose per-table period filtering; check the XLSX itself."
        ),
        max_length=40,
    )


class LookupCommuneResult(BaseModel):
    source: str = AGVCH_ATTRIBUTION
    provenance: str | None = None
    error: str | None = None
    hint: str | None = None
    query: str | None = None
    valid_at_date: str | None = None
    total_matches: int | None = None
    communes: list[CommuneEntry] | None = None
    note: str | None = None


class ResolveHistoricalCommuneResult(BaseModel):
    source: str = AGVCH_ATTRIBUTION
    provenance: str | None = None
    error: str | None = None
    hint: str | None = None
    bfs_number: int | None = None
    from_date: str | None = None
    to_date: str | None = None
    unchanged: bool | None = None
    resolves_to: list[SuccessorEntry] | None = None  # today's BFS number(s)
    mutation_path: list[MutationStep] | None = None
    note: str | None = None


class ListCommunesResult(BaseModel):
    source: str = AGVCH_ATTRIBUTION
    provenance: str | None = None
    error: str | None = None
    hint: str | None = None
    canton: str | None = None
    canton_abbr: str | None = None
    valid_at_date: str | None = None
    total: int | None = None
    communes: list[CommuneEntry] | None = None
    note: str | None = None


class SearchHistoricalSeriesResult(BaseModel):
    source: str = HSSO_ATTRIBUTION
    provenance: str | None = None
    licence_note: str = HSSO_NONCOMMERCIAL_NOTE
    error: str | None = None
    hint: str | None = None
    topic: str | None = None
    period: str | None = None
    total_matches: int | None = None
    series: list[HistoricalSeriesEntry] | None = None
    note: str | None = None


# ---------------------------------------------------------------------------
# Response formatting helpers
# ---------------------------------------------------------------------------

def _format_jsonstat2_as_table(data: dict[str, Any], max_rows: int = 500) -> dict[str, Any]:
    """Convert JSON-stat2 response to a readable table format."""
    dimensions = data.get("id", [])
    dimension_info = data.get("dimension", {})
    values = data.get("value", [])
    size = data.get("size", [])

    # Build label lookups for each dimension
    dim_labels: list[list[tuple[str, str]]] = []
    for dim_id in dimensions:
        dim_data = dimension_info.get(dim_id, {})
        cats = dim_data.get("category", {})
        index_map = cats.get("index", {})
        label_map = cats.get("label", {})

        if isinstance(index_map, dict):
            # index_map: {value_code: position}
            ordered = sorted(index_map.items(), key=lambda x: x[1])
            pairs = [(code, label_map.get(code, code)) for code, _ in ordered]
        else:
            # index_map is a list
            pairs = [(code, label_map.get(code, code)) for code in index_map]

        if not pairs:
            # Fallback: use label map directly
            pairs = list(label_map.items())

        dim_labels.append(pairs)

    # Generate all combinations
    import itertools
    combos = list(itertools.product(*dim_labels))

    rows = []
    for i, (combo, value) in enumerate(zip(combos, values)):
        if i >= max_rows:
            break
        row: dict[str, str | float | None] = {}
        for dim_id, (code, label) in zip(dimensions, combo):
            row[dim_id] = label
        row["Wert"] = value
        rows.append(row)

    rows_total = len(values)
    rows_returned = len(rows)
    return {
        "title": data.get("label", ""),
        "source": data.get("source", "BFS"),
        "updated": data.get("updated", ""),
        "dimensions": [
            {
                "id": dim_id,
                "label": dimension_info.get(dim_id, {}).get("label", dim_id),
                "n_values": sz,
            }
            for dim_id, sz in zip(dimensions, size)
        ],
        "rows_total": rows_total,
        "rows_returned": rows_returned,
        "truncated": rows_returned < rows_total,
        "rows": rows,
    }


# ---------------------------------------------------------------------------
# MCP Server
# ---------------------------------------------------------------------------

mcp = FastMCP(
    "swiss_statistics_mcp",
    instructions=(
        "Access Swiss Federal Statistical Office (BFS/OFS/UST) data via STAT-TAB. "
        "Available: 682 datasets across 21 themes. No API key required. "
        "Workflow: (1) bfs_list_themes to see themes, (2) bfs_search_tables or "
        "bfs_list_tables_by_theme to find datasets, (3) bfs_get_table_metadata to "
        "understand variables and valid filter values, (4) bfs_get_data to retrieve "
        "actual statistics. Use bfs_education_stats for Schulamt-relevant shortcuts. "
        "Reference layer: lookup_commune / list_communes / resolve_historical_commune "
        "resolve official BFS commune numbers (the portfolio join key) and re-key old "
        "statistics across fusions; search_historical_series finds long-run HSSO series. "
        "Construction (STAT-TAB theme 09): bfs_construction_activity for new "
        "buildings/dwellings per commune, bfs_construction_investment for building "
        "investment and Arbeitsvorrat (the monetary leading indicator) by region/canton/commune. "
        "Price indices (not in STAT-TAB): bfs_price_index for the construction price "
        "index (Baupreisindex, parsed series) and the residential property price index "
        "(IMPI, source links only — BFS publishes it as PDF)."
    ),
)


# ---------------------------------------------------------------------------
# Tool: List themes
# ---------------------------------------------------------------------------

@mcp.tool(
    name="bfs_list_themes",
    annotations={
        "title": "BFS Statistical Themes",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
@_logged_tool("bfs_list_themes")
async def bfs_list_themes(params: ListThemesInput) -> ListThemesResult:
    """List all 21 BFS statistical themes with their codes and dataset counts.

    Returns the complete taxonomy of Swiss federal statistics. Each theme has
    a 2-digit code used to filter datasets with bfs_list_tables_by_theme.

    Args:
        params (ListThemesInput):
            - lang (str): Language code ('de', 'fr', 'it', 'en')

    Returns:
        ListThemesResult with theme codes, names, dataset counts per theme.
        On error, `error` and `hint` are set and `themes` is None.
    """
    try:
        url = f"{BFS_API_BASE}/{params.lang}/"
        all_dbs = await _get(url)

        # Count tables per theme
        theme_counts: dict[str, int] = {code: 0 for code in BFS_THEMES}
        for db in all_dbs:
            code = _theme_code_from_dbid(db["dbid"])
            if code in theme_counts:
                theme_counts[code] += 1

        themes = [
            ThemeEntry(
                code=code,
                name=name,
                dataset_count=theme_counts.get(code, 0),
                filter_hint=f"Use theme_code='{code}' in bfs_list_tables_by_theme",
            )
            for code, name in BFS_THEMES.items()
        ]

        return ListThemesResult(
            total_datasets=len(all_dbs),
            themes=themes,
            note=(
                "Use bfs_list_tables_by_theme(theme_code='15') for Bildung, "
                "bfs_search_tables(query='Lehrpersonen') for keyword search."
            ),
        )
    except httpx.HTTPStatusError as e:
        return ListThemesResult(
            error=f"API-Fehler {e.response.status_code}",
            hint="BFS STAT-TAB API nicht erreichbar. Bitte später nochmals versuchen.",
        )
    except Exception:
        _LOGGER.exception("bfs_list_themes failed")
        return ListThemesResult(
            error="Interner Fehler beim Laden der Themen.",
            hint="Bitte erneut versuchen.",
        )


# ---------------------------------------------------------------------------
# Tool: List tables by theme
# ---------------------------------------------------------------------------

@mcp.tool(
    name="bfs_list_tables_by_theme",
    annotations={
        "title": "BFS Tables by Theme",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
@_logged_tool("bfs_list_tables_by_theme")
async def bfs_list_tables_by_theme(params: ListTablesByThemeInput) -> ListTablesByThemeResult:
    """List available statistical tables for a specific BFS theme.

    Returns table IDs and titles for a given theme code. Use the returned
    table_id values with bfs_get_table_metadata and bfs_get_data.

    Args:
        params (ListTablesByThemeInput):
            - theme_code (str): 2-digit theme code, e.g. '15' for Bildung
            - lang (str): Language code
            - limit (int): Max tables to return (default 20)

    Returns:
        ListTablesByThemeResult with the matching `tables` list. On error,
        `error` and `hint` are set and `tables` is None.
    """
    try:
        url = f"{BFS_API_BASE}/{params.lang}/"
        all_dbs = await _get(url)

        # Filter by theme
        theme_dbs = [
            db for db in all_dbs
            if _theme_code_from_dbid(db["dbid"]) == params.theme_code
        ]

        if not theme_dbs:
            available = list(BFS_THEMES.keys())
            return ListTablesByThemeResult(
                error=f"Kein Thema mit Code '{params.theme_code}' gefunden.",
                hint=f"Verfügbare Codes: {available}. Verwende bfs_list_themes für die vollständige Liste.",
            )

        theme_name = BFS_THEMES.get(params.theme_code, params.theme_code)

        # Fan-out metadata fetches with a Semaphore to avoid hammering BFS.
        # Each fetch goes through `_fetch_metadata_cached`, so warm calls
        # are O(1) and only cold IDs hit the wire. (SCALE-003, SCALE-004)
        sem = asyncio.Semaphore(FANOUT_CONCURRENCY)
        selected = [db["dbid"] for db in theme_dbs[: params.limit]]

        async def fetch_one(dbid: str) -> TableEntry:
            async with sem:
                try:
                    meta = await _fetch_metadata_cached(dbid, params.lang)
                    return TableEntry(
                        table_id=dbid,
                        title=meta.get("title", dbid),
                        last_updated=meta.get("updated", ""),
                        n_variables=len(meta.get("variables", [])),
                        featured=dbid in FEATURED_TABLES,
                    )
                except Exception:
                    _LOGGER.warning(
                        "table metadata fetch failed for %s", dbid, exc_info=True
                    )
                    return TableEntry(table_id=dbid, title=dbid)

        tables = await asyncio.gather(*(fetch_one(dbid) for dbid in selected))

        return ListTablesByThemeResult(
            theme_code=params.theme_code,
            theme_name=theme_name,
            total_in_theme=len(theme_dbs),
            returned=len(tables),
            tables=list(tables),
            next_step=(
                "Verwende bfs_get_table_metadata(table_id='...') "
                "um Variablen und Filter-Werte zu sehen."
            ),
        )
    except httpx.HTTPStatusError as e:
        return ListTablesByThemeResult(error=f"API-Fehler {e.response.status_code}")
    except Exception:
        _LOGGER.exception("bfs_list_tables_by_theme failed")
        return ListTablesByThemeResult(
            error="Interner Fehler beim Laden der Tabellen-Liste.",
            hint="Bitte erneut versuchen.",
        )


# ---------------------------------------------------------------------------
# Tool: Search tables
# ---------------------------------------------------------------------------

@mcp.tool(
    name="bfs_search_tables",
    annotations={
        "title": "Search BFS Tables",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": False,
        "openWorldHint": False,
    },
)
@_logged_tool("bfs_search_tables")
async def bfs_search_tables(params: SearchTablesInput) -> SearchTablesResult:
    """Search for BFS statistical tables by keyword in their titles.

    Performs a full-text search across all 682+ BFS table titles.
    Results include table IDs needed for bfs_get_table_metadata and bfs_get_data.

    Note: First call builds a catalog (~682 API requests). Subsequent calls
    within 1 hour use the cached catalog and are instant.

    Args:
        params (SearchTablesInput):
            - query (str): Search keywords, e.g. 'Lehrkräfte', 'Schüler Kanton'
            - theme_code (Optional[str]): Filter by theme, e.g. '15'
            - lang (str): Language for table titles
            - limit (int): Max results (default 10)

    Returns:
        SearchTablesResult with matching tables. On error, `error` and
        `hint` are set and `results` is None.
    """
    try:
        catalog = await _ensure_catalog(params.lang)

        query_lower = params.query.lower()
        query_terms = query_lower.split()

        results: list[SearchResultEntry] = []
        for dbid, title in catalog.items():
            # Filter by theme if specified
            if params.theme_code:
                if _theme_code_from_dbid(dbid) != params.theme_code:
                    continue

            # Match all query terms in title
            title_lower = title.lower()
            if all(term in title_lower for term in query_terms):
                theme_code = _theme_code_from_dbid(dbid)
                results.append(
                    SearchResultEntry(
                        table_id=dbid,
                        title=title,
                        theme_code=theme_code,
                        theme_name=BFS_THEMES.get(theme_code, theme_code),
                        featured=dbid in FEATURED_TABLES,
                    )
                )

        results = results[: params.limit]

        return SearchTablesResult(
            query=params.query,
            total_matches=len(results),
            results=results,
            next_step=(
                "Verwende bfs_get_table_metadata(table_id='...') "
                "um die Variablen einer Tabelle zu sehen."
            ),
        )
    except httpx.HTTPStatusError as e:
        return SearchTablesResult(error=f"API-Fehler {e.response.status_code}")
    except Exception:
        _LOGGER.exception("bfs_search_tables failed")
        return SearchTablesResult(
            error="Interner Fehler beim Aufbau des Katalogs.",
            hint="Bitte erneut versuchen.",
        )


# ---------------------------------------------------------------------------
# Tool: Get table metadata
# ---------------------------------------------------------------------------

@mcp.tool(
    name="bfs_get_table_metadata",
    annotations={
        "title": "BFS Table Metadata",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
@_logged_tool("bfs_get_table_metadata")
async def bfs_get_table_metadata(params: GetTableMetadataInput) -> TableMetadataResult:
    """Get metadata for a BFS table: title, variables, and available filter values.

    Essential step before calling bfs_get_data. Returns all dimension variables
    with their codes and value labels needed to construct data queries.

    Args:
        params (GetTableMetadataInput):
            - table_id (str): BFS table ID, e.g. 'px-x-1504000000_173'
            - lang (str): Language for labels

    Returns:
        str: JSON with table title, source, update date, and all variables with
             their codes and value options. Use variable codes in bfs_get_data filters.

    Example output structure:
        {
          "title": "Lehrkräfte nach Schuljahr, Kanton...",
          "variables": [
            {
              "code": "Schuljahr",
              "label": "Schuljahr",
              "n_values": 14,
              "values": [{"code": "0", "label": "2010/11"}, ...]
            }
          ]
        }
    """
    try:
        meta = await _fetch_metadata_cached(params.table_id, params.lang)

        variables: list[VariableInfo] = []
        for var in meta.get("variables", []):
            value_pairs = [
                VariableValue(code=code, label=label)
                for code, label in zip(
                    var.get("values", []),
                    var.get("valueTexts", var.get("values", [])),
                )
            ]
            variables.append(
                VariableInfo(
                    code=var.get("code", ""),
                    label=var.get("text", ""),
                    n_values=len(value_pairs),
                    values=value_pairs[:30],  # cap displayed values
                    more_values=max(0, len(value_pairs) - 30),
                )
            )

        theme_code = _theme_code_from_dbid(params.table_id)

        first_var_code = variables[0].code if variables else "Variable"
        first_val_code = (
            variables[0].values[0].code
            if variables and variables[0].values
            else "0"
        )

        return TableMetadataResult(
            table_id=params.table_id,
            title=meta.get("title", ""),
            source=meta.get("source", "BFS"),
            last_updated=meta.get("updated", ""),
            theme_code=theme_code,
            theme_name=BFS_THEMES.get(theme_code, ""),
            language=params.lang,
            n_variables=len(variables),
            variables=variables,
            usage_hint=(
                "Verwende 'code' der Variable und 'code' der gewünschten Werte "
                "als Filter in bfs_get_data. Beispiel: "
                f"filters=[{{\"code\": \"{first_var_code}\", "
                f"\"values\": [\"{first_val_code}\"]}}]"
            ),
        )
    except httpx.HTTPStatusError as e:
        if e.response.status_code == 404:
            return TableMetadataResult(
                error=f"Tabelle '{params.table_id}' nicht gefunden.",
                hint="Verwende bfs_search_tables oder bfs_list_tables_by_theme um gültige IDs zu finden.",
            )
        return TableMetadataResult(error=f"API-Fehler {e.response.status_code}")
    except Exception:
        _LOGGER.exception("bfs_get_table_metadata failed")
        return TableMetadataResult(
            error="Interner Fehler beim Laden der Metadaten.",
            hint="Bitte erneut versuchen.",
        )


# ---------------------------------------------------------------------------
# Tool: Get data
# ---------------------------------------------------------------------------

@mcp.tool(
    name="bfs_get_data",
    annotations={
        "title": "Get BFS Statistical Data",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
@_logged_tool("bfs_get_data")
async def bfs_get_data(params: GetDataInput) -> DataTableResult:
    """Query statistical data from a BFS table with optional filters.

    Fetches actual data values from a STAT-TAB table. Always call
    bfs_get_table_metadata first to understand available variables and values.

    Args:
        params (GetDataInput):
            - table_id (str): BFS table ID
            - filters (Optional[list]): Dimension filters to narrow results.
              Each filter: {"code": "VariableCode", "values": ["val1", "val2"]}
              Without filters, all data is returned (may be very large).
            - lang (str): Language for labels
            - max_rows (int): Safety limit on returned rows (default 500)

    Returns:
        DataTableResult with `dimensions`, `rows`, plus `truncated`,
        `rows_total`, `rows_returned` for machine-readable capping.
        On error, `error` and `hint` are set.
    """
    try:
        url = _build_data_url(params.table_id, params.lang)

        # Build PxWeb query
        query: list[dict[str, Any]] = []
        if params.filters:
            for f in params.filters:
                query.append(
                    {
                        "code": f.code,
                        "selection": {"filter": "item", "values": f.values},
                    }
                )

        body = {"query": query, "response": {"format": "json-stat2"}}
        data = await _post(url, body)

        formatted = _format_jsonstat2_as_table(data, max_rows=params.max_rows)
        note = None
        if formatted["truncated"]:
            note = (
                f"Datenmenge auf {params.max_rows} Zeilen begrenzt "
                f"(total: {formatted['rows_total']}). "
                "Verwende Filter um die Datenmenge einzuschränken."
            )

        return DataTableResult(
            table_id=params.table_id,
            language=params.lang,
            note=note,
            **formatted,
        )

    except httpx.HTTPStatusError as e:
        if e.response.status_code == 404:
            return DataTableResult(
                error=f"Tabelle '{params.table_id}' nicht gefunden.",
                hint="Prüfe die table_id mit bfs_search_tables.",
            )
        if e.response.status_code == 400:
            return DataTableResult(
                error="Ungültige Abfrage (HTTP 400).",
                hint=(
                    "Prüfe ob die Filter-Codes und -Werte korrekt sind. "
                    "Verwende bfs_get_table_metadata um gültige Codes zu erhalten."
                ),
            )
        return DataTableResult(
            error=f"API-Fehler {e.response.status_code}: {e.response.text[:200]}"
        )
    except Exception:
        _LOGGER.exception("bfs_get_data failed")
        return DataTableResult(
            error="Interner Fehler beim Daten-Abruf.",
            hint="Bitte mit kleinerer Datenmenge oder anderen Filtern erneut versuchen.",
        )


# ---------------------------------------------------------------------------
# Tool: Education stats (Schulamt convenience)
# ---------------------------------------------------------------------------

EDUCATION_TOPIC_MAP: dict[str, dict[str, Any]] = {
    "teachers": {
        "table_id": "px-x-1504000000_173",
        "description": "Lehrkräfte nach Schuljahr, Kanton, Beschäftigungsgrad und Bildungsstufe",
        "canton_var": "Kanton",
        "canton_all_value": "0",
    },
    "students": {
        "table_id": "px-x-1502020100_101",
        "description": "Schülerinnen und Schüler nach Bildungsstufe und Kanton",
        "canton_var": "Kanton",
        "canton_all_value": "0",
    },
    "scenarios": {
        "table_id": "px-x-1509090000_101",
        "description": "Szenarien 2022-2031: Entwicklung Schülerzahlen Sekundarstufe II",
        "canton_var": "Kanton",
        "canton_all_value": "0",
    },
    "scholarships": {
        "table_id": "px-x-1506020000_114",
        "description": "Stipendien und Darlehen nach Kanton",
        "canton_var": "Kanton",
        "canton_all_value": "0",
    },
}

CANTON_NAME_TO_VALUE: dict[str, str] = {
    "Schweiz": "0",
    "Zürich": "1",
    "Bern / Berne": "2",
    "Luzern": "3",
    "Uri": "4",
    "Schwyz": "5",
    "Obwalden": "6",
    "Nidwalden": "7",
    "Glarus": "8",
    "Zug": "9",
    "Freiburg / Fribourg": "10",
    "Solothurn": "11",
    "Basel-Stadt": "12",
    "Basel-Landschaft": "13",
    "Schaffhausen": "14",
    "Appenzell Ausserrhoden": "15",
    "Appenzell Innerrhoden": "16",
    "St. Gallen": "17",
    "Graubünden / Grigioni / Grischun": "18",
    "Aargau": "19",
    "Thurgau": "20",
    "Ticino": "21",
    "Vaud": "22",
    "Valais / Wallis": "23",
    "Neuchâtel": "24",
    "Genève": "25",
    "Jura": "26",
}

# Population table uses cantonal abbreviations (BFS codes), not numeric indices
CANTON_POPULATION_CODE: dict[str, str] = {
    "Schweiz": "8100",
    "Zürich": "ZH",
    "Bern / Berne": "BE",
    "Luzern": "LU",
    "Uri": "UR",
    "Schwyz": "SZ",
    "Obwalden": "OW",
    "Nidwalden": "NW",
    "Glarus": "GL",
    "Zug": "ZG",
    "Fribourg / Freiburg": "FR",
    "Solothurn": "SO",
    "Basel-Stadt": "BS",
    "Basel-Landschaft": "BL",
    "Schaffhausen": "SH",
    "Appenzell Ausserrhoden": "AR",
    "Appenzell Innerrhoden": "AI",
    "St. Gallen": "SG",
    "Graubünden / Grigioni / Grischun": "GR",
    "Aargau": "AG",
    "Thurgau": "TG",
    "Ticino": "TI",
    "Vaud": "VD",
    "Valais / Wallis": "VS",
    "Neuchâtel": "NE",
    "Genève": "GE",
    "Jura": "JU",
}


@mcp.tool(
    name="bfs_education_stats",
    annotations={
        "title": "Swiss Education Statistics",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
@_logged_tool("bfs_education_stats")
async def bfs_education_stats(params: GetEducationStatsInput) -> DataTableResult:
    """Retrieve Swiss education statistics — convenience tool for Schulamt context.

    Provides direct access to key education datasets without needing to know
    table IDs or variable codes. Covers teachers, students, enrollment scenarios,
    and scholarship data, optionally filtered by canton.

    Args:
        params (GetEducationStatsInput):
            - topic (str): One of: 'teachers', 'students', 'scenarios', 'scholarships'
            - canton (Optional[str]): Canton name, e.g. 'Zürich'. None = all cantons.
            - lang (str): Language code

    Returns:
        DataTableResult with `topic`, `topic_description`, `canton_filter`
        on success, plus the data table fields. On error, `error` and
        `hint` are set.
    """
    topic_cfg = EDUCATION_TOPIC_MAP[params.topic]
    table_id: str = topic_cfg["table_id"]
    canton_var: str = topic_cfg["canton_var"]

    # Resolve canton to value code
    canton_value: str | None = None
    if params.canton:
        # Try exact match first, then partial match
        canton_value = CANTON_NAME_TO_VALUE.get(params.canton)
        if canton_value is None:
            for name, val in CANTON_NAME_TO_VALUE.items():
                if params.canton.lower() in name.lower():
                    canton_value = val
                    break
        if canton_value is None:
            return DataTableResult(
                error=f"Kanton '{params.canton}' nicht gefunden.",
                hint=f"Gültige Kantone: {list(CANTON_NAME_TO_VALUE.keys())}",
            )

    try:
        url = _build_data_url(table_id, params.lang)

        query: list[dict[str, Any]] = []
        if canton_value is not None:
            query.append(
                {
                    "code": canton_var,
                    "selection": {"filter": "item", "values": [canton_value]},
                }
            )

        body = {"query": query, "response": {"format": "json-stat2"}}
        data = await _post(url, body)

        formatted = _format_jsonstat2_as_table(data, max_rows=500)

        return DataTableResult(
            table_id=table_id,
            topic=params.topic,
            topic_description=topic_cfg["description"],
            canton_filter=params.canton,
            language=params.lang,
            **formatted,
        )

    except httpx.HTTPStatusError as e:
        return DataTableResult(
            error=f"API-Fehler {e.response.status_code}",
            hint="Tabelle möglicherweise aktuell nicht verfügbar. Bitte später nochmals versuchen.",
        )
    except Exception:
        _LOGGER.exception("bfs_education_stats failed")
        return DataTableResult(
            error="Interner Fehler beim Abruf der Bildungsstatistiken.",
            hint="Bitte erneut versuchen.",
        )


# ---------------------------------------------------------------------------
# Tool: Population data
# ---------------------------------------------------------------------------

@mcp.tool(
    name="bfs_population",
    annotations={
        "title": "Swiss Population Statistics",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
@_logged_tool("bfs_population")
async def bfs_population(params: GetPopulationInput) -> DataTableResult:
    """Retrieve Swiss population statistics by region, year, and breakdown.

    Accesses the core BFS population dataset (ständige Wohnbevölkerung)
    with flexible filtering by canton/municipality, year, age, and gender.
    Critical for school space planning and demographic projections.

    Args:
        params (GetPopulationInput):
            - region (str): 'Schweiz', or canton name like 'Zürich'
            - year (Optional[str]): Year filter, e.g. '2024'
            - breakdown (str): 'total', 'age', or 'gender'

    Returns:
        str: JSON with population figures for the selected region and breakdown.
    """
    TABLE_ID = "px-x-0102010000_101"

    # Map region to canton code (population table uses BFS abbreviations: ZH, BE, etc.)
    region_value = CANTON_POPULATION_CODE.get(params.region)
    if region_value is None:
        for name, val in CANTON_POPULATION_CODE.items():
            if params.region.lower() in name.lower():
                region_value = val
                break

    if region_value is None:
        return DataTableResult(
            error=f"Region '{params.region}' nicht gefunden.",
            hint="Verwende 'Schweiz' oder einen Kantonnamen wie 'Zürich', 'Bern / Berne'.",
        )

    canton_code = region_value

    try:
        url = _build_data_url(TABLE_ID, "de")

        # Build filters based on breakdown
        query: list[dict[str, Any]] = [
            {
                "code": "Kanton (-) / Bezirk (>>) / Gemeinde (......)",
                "selection": {"filter": "item", "values": [canton_code]},
            },
            {
                "code": "Bevölkerungstyp",
                "selection": {"filter": "item", "values": ["1"]},  # Ständige Wohnbevölkerung
            },
        ]

        # Nationality: total
        query.append(
            {
                "code": "Staatsangehörigkeit (Kategorie)",
                "selection": {"filter": "item", "values": ["-99999"]},  # Total
            }
        )

        if params.breakdown == "total":
            query.append(
                {"code": "Geschlecht", "selection": {"filter": "item", "values": ["-99999"]}}
            )
            query.append(
                {"code": "Alter", "selection": {"filter": "item", "values": ["-99999"]}}
            )
        elif params.breakdown == "gender":
            query.append(
                {"code": "Geschlecht", "selection": {"filter": "item", "values": ["1", "2"]}}
            )
            query.append(
                {"code": "Alter", "selection": {"filter": "item", "values": ["-99999"]}}
            )
        elif params.breakdown == "age":
            query.append(
                {"code": "Geschlecht", "selection": {"filter": "item", "values": ["-99999"]}}
            )
            # Age groups: 0-18 for school planning context
            age_values = [str(i) for i in range(19)]
            query.append(
                {"code": "Alter", "selection": {"filter": "item", "values": age_values}}
            )

        if params.year:
            query.append(
                {"code": "Jahr", "selection": {"filter": "item", "values": [params.year]}}
            )

        body = {"query": query, "response": {"format": "json-stat2"}}
        data = await _post(url, body)

        formatted = _format_jsonstat2_as_table(data, max_rows=200)

        return DataTableResult(
            table_id=TABLE_ID,
            region=params.region,
            breakdown=params.breakdown,
            year=params.year,
            language="de",
            note=(
                "Ständige Wohnbevölkerung. "
                "Für Schulraumplanung empfiehlt sich breakdown='age' für Altersgruppen 0-18."
            ),
            **formatted,
        )

    except httpx.HTTPStatusError as e:
        return DataTableResult(error=f"API-Fehler {e.response.status_code}")
    except Exception:
        _LOGGER.exception("bfs_population failed")
        return DataTableResult(
            error="Interner Fehler beim Abruf der Bevölkerungsdaten.",
            hint="Bitte erneut versuchen.",
        )


# ---------------------------------------------------------------------------
# Tool: Compare cantons
# ---------------------------------------------------------------------------

@mcp.tool(
    name="bfs_compare_cantons",
    annotations={
        "title": "Compare Swiss Cantons",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
@_logged_tool("bfs_compare_cantons")
async def bfs_compare_cantons(params: CompareCantonsInput) -> DataTableResult:
    """Compare a BFS statistical indicator across multiple Swiss cantons.

    Designed for KI-Fachgruppe demos and benchmarking. Fetches the same
    dataset for multiple cantons simultaneously, enabling direct comparison.

    Args:
        params (CompareCantonsInput):
            - table_id (str): BFS table ID to query
            - canton_values (list[str]): Canton value codes to compare.
              Use '0' for Switzerland total, '1' for Zürich, '2' for Bern, etc.
              Get codes via bfs_get_table_metadata on any canton-level table.
            - additional_filters (Optional[list]): Extra dimension filters
            - lang (str): Language code

    Returns:
        str: JSON with data for all selected cantons side by side.

    Example use case:
        Compare teacher-to-student ratios across ZH, BE, LU, CH total:
        canton_values=['0', '1', '2', '3']
    """
    try:
        url = _build_data_url(params.table_id, params.lang)

        # Build query with canton filter
        query: list[dict[str, Any]] = []

        # Find canton variable name from metadata first (cache-hit if any
        # previous tool resolved this table)
        meta = await _fetch_metadata_cached(params.table_id, params.lang)

        canton_var_code = None
        for var in meta.get("variables", []):
            var_code = var.get("code", "").lower()
            if "kanton" in var_code:
                canton_var_code = var["code"]
                break

        if canton_var_code is None:
            return DataTableResult(
                error="Keine Kanton-Variable in dieser Tabelle gefunden.",
                hint=f"Verfügbare Variablen: {[v['code'] for v in meta.get('variables', [])]}",
            )

        query.append(
            {
                "code": canton_var_code,
                "selection": {"filter": "item", "values": params.canton_values},
            }
        )

        if params.additional_filters:
            for f in params.additional_filters:
                query.append(
                    {
                        "code": f.code,
                        "selection": {"filter": "item", "values": f.values},
                    }
                )

        body = {"query": query, "response": {"format": "json-stat2"}}
        data = await _post(url, body)

        formatted = _format_jsonstat2_as_table(data, max_rows=500)

        return DataTableResult(
            table_id=params.table_id,
            language=params.lang,
            cantons_compared=params.canton_values,
            canton_variable=canton_var_code,
            **formatted,
        )

    except httpx.HTTPStatusError as e:
        if e.response.status_code == 400:
            return DataTableResult(
                error="Ungültige Anfrage. Prüfe ob die Kanton-Werte für diese Tabelle gültig sind.",
                hint="Verwende bfs_get_table_metadata um gültige Werte-Codes zu erhalten.",
            )
        return DataTableResult(error=f"API-Fehler {e.response.status_code}")
    except Exception:
        _LOGGER.exception("bfs_compare_cantons failed")
        return DataTableResult(
            error="Interner Fehler beim Kantons-Vergleich.",
            hint="Bitte erneut versuchen.",
        )


# ---------------------------------------------------------------------------
# Tool: Featured datasets
# ---------------------------------------------------------------------------

@mcp.tool(
    name="bfs_featured_datasets",
    annotations={
        "title": "BFS Featured Datasets",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
@_logged_tool("bfs_featured_datasets")
async def bfs_featured_datasets(params: ListThemesInput) -> FeaturedDatasetsResult:
    """Return a curated list of high-value BFS datasets for Schulamt and public administration.

    Provides a shortlist of the most relevant datasets for education planning,
    demographic analysis, and political context — ideal as a starting point.

    Args:
        params (ListThemesInput):
            - lang (str): Language code

    Returns:
        FeaturedDatasetsResult with curated table IDs, titles, themes, and
        recommended use cases.
    """
    featured = [
        FeaturedDatasetEntry(
            table_id=tid,
            title=FEATURED_TABLES[tid],
            theme_code=_theme_code_from_dbid(tid),
            theme_name=BFS_THEMES.get(_theme_code_from_dbid(tid), ""),
            schulamt_relevanz=_schulamt_relevance(tid),
        )
        for tid in FEATURED_TABLES
    ]

    return FeaturedDatasetsResult(
        total=len(featured),
        featured_datasets=featured,
        quick_start=(
            "Verwende bfs_education_stats(topic='teachers') für Lehrkräfte-Statistiken, "
            "bfs_population(region='Zürich', breakdown='age') für Altersstruktur in Zürich, "
            "oder bfs_get_table_metadata(table_id='px-x-1504000000_173') "
            "für detaillierte Lehrkräfte-Daten."
        ),
    )


def _schulamt_relevance(table_id: str) -> str:
    relevance_map = {
        "px-x-1504000000_173": "⭐⭐⭐ Kerndaten Lehrpersonenmangel – nach Kanton & Bildungsstufe",
        "px-x-1504000000_172": "⭐⭐⭐ Lehrpersonen nach Staatsangehörigkeit – Diversity-Analyse",
        "px-x-0102010000_101": "⭐⭐⭐ Schulraumplanung – Bevölkerung nach Alter & Gemeinde",
        "px-x-1509090000_101": "⭐⭐⭐ Prognosen Schülerzahlen Sek II bis 2031",
        "px-x-1509090000_113": "⭐⭐ Prognosen Hochschulen – Lehrpersonen-Nachwuchs",
        "px-x-1502020100_101": "⭐⭐⭐ Schülerbestände nach Bildungsstufe – Kantonsvergleich",
        "px-x-1503040100_101": "⭐⭐ Abschlüsse Sek II – Bildungsoutput-Analyse",
        "px-x-1506020000_114": "⭐⭐ Stipendien – Bildungsfinanzierung nach Kanton",
        "px-x-1703030000_101": "⭐ Nationalratswahlen – politischer Kontext",
        "px-x-1703030000_100": "⭐ Volksabstimmungen – demokratische Legitimation",
        "px-x-0301000000_101": "⭐⭐ Arbeitsmarkt – Rahmenbedingungen Personalrekrutierung",
        "px-x-1302020000_101": "⭐⭐ Sozialhilfe – sozialer Kontext Volksschule",
    }
    return relevance_map.get(table_id, "Relevant für öffentliche Verwaltung")


# ---------------------------------------------------------------------------
# Reference layer tools — AGVCH commune register (Architecture A)
# ---------------------------------------------------------------------------


@mcp.tool(
    name="lookup_commune",
    annotations={
        "title": "Look up a Swiss commune",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
@_logged_tool("lookup_commune")
async def lookup_commune(params: LookupCommuneInput) -> LookupCommuneResult:
    """Resolve a Swiss commune by name or BFS number, as of a given date.

    The BFS commune number is the portfolio's join key. This tool returns
    the official register entry — BFS number, name, canton, validity dates
    and the stable LINDAS URI — for a commune as it existed on `valid_at_date`.

    Args:
        params (LookupCommuneInput):
            - name_or_bfs_number (str): name/substring or BFS number
            - valid_at_date (str): ISO date; commune state as of this date

    Returns:
        LookupCommuneResult with matching `communes` (BFS number, canton,
        validity, LINDAS URI). On error, `error` and `hint` are set.
    """
    try:
        agvch_date = _iso_to_agvch(params.valid_at_date)
        rows, from_cache = await _fetch_snapshot(agvch_date)
        by_hist = _index_by_hist(rows)

        query = params.name_or_bfs_number.strip()
        is_number = query.isdigit()
        query_lower = query.lower()

        matches: list[dict[str, str]] = []
        for r in rows:
            if r.get("Level") != "3":  # communes only
                continue
            if is_number:
                if r.get("BfsCode") == query:
                    matches.append(r)
            elif query_lower in r.get("Name", "").lower():
                matches.append(r)

        provenance = "cached" if from_cache else "live_api"
        if not matches:
            return LookupCommuneResult(
                provenance=provenance,
                query=query,
                valid_at_date=params.valid_at_date,
                total_matches=0,
                communes=[],
                note=(
                    f"Keine Gemeinde für '{query}' zum {params.valid_at_date} gefunden. "
                    "Bei historischen Nummern resolve_historical_commune verwenden."
                ),
            )

        entries = [_commune_entry(r, by_hist) for r in matches[:50]]
        entries.sort(key=lambda e: e.bfs_number)
        return LookupCommuneResult(
            provenance=provenance,
            query=query,
            valid_at_date=params.valid_at_date,
            total_matches=len(matches),
            communes=entries,
            note=(
                f"{len(matches)} Treffer (angezeigt: {len(entries)})."
                if len(matches) > len(entries)
                else None
            ),
        )
    except httpx.HTTPStatusError as e:
        return LookupCommuneResult(
            error=f"AGVCH-API-Fehler {e.response.status_code}",
            hint="Datum im Format YYYY-MM-DD prüfen. Quelle evtl. kurz nicht erreichbar.",
        )
    except Exception:
        _LOGGER.exception("lookup_commune failed")
        return LookupCommuneResult(
            error="Interner Fehler beim Gemeinde-Lookup.",
            hint="Bitte erneut versuchen.",
        )


@mcp.tool(
    name="resolve_historical_commune",
    annotations={
        "title": "Resolve a historical commune to today's BFS number",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
@_logged_tool("resolve_historical_commune")
async def resolve_historical_commune(
    params: ResolveHistoricalCommuneInput,
) -> ResolveHistoricalCommuneResult:
    """Map a historical BFS commune number onto today's number(s).

    This is the core value of the reference layer: when old statistics are
    keyed on a BFS number that has since been merged or renamed, this tool
    returns which of today's commune(s) that number resolves to, plus the
    mutation path (fusions/renamings with dates). Use `resolves_to` to
    re-key (umschlüsseln) old figures onto the current municipal division.

    Args:
        params (ResolveHistoricalCommuneInput):
            - bfs_number (int): historical BFS number
            - from_date (str): ISO date the old data belongs to
            - to_date (str): ISO target date (default today)

    Returns:
        ResolveHistoricalCommuneResult with `resolves_to` (today's BFS
        number/name/LINDAS URI) and `mutation_path`. On error, `error`/`hint`.
    """
    try:
        start = _iso_to_agvch(params.from_date)
        end = _iso_to_agvch(params.to_date)
        bfs = str(params.bfs_number)

        corr = await _fetch_agvch_csv(
            "correspondances",
            {
                "includeUnmodified": "true",
                "includeTerritoryExchange": "false",
                "startPeriod": start,
                "endPeriod": end,
            },
        )
        related = [row for row in corr if row.get("InitialCode") == bfs]
        if not related:
            return ResolveHistoricalCommuneResult(
                provenance="live_api",
                bfs_number=params.bfs_number,
                from_date=params.from_date,
                to_date=params.to_date,
                error=f"Keine Gemeinde mit BFS-Nummer {bfs} zum {params.from_date} gefunden.",
                hint=(
                    "Die Gemeinde muss zum from_date existiert haben. "
                    "BFS-Nummer und Startdatum prüfen (lookup_commune)."
                ),
            )

        successors: list[SuccessorEntry] = []
        seen: set[str] = set()
        for row in related:
            tc = row.get("TerminalCode") or ""
            if tc and tc not in seen:
                seen.add(tc)
                successors.append(
                    SuccessorEntry(
                        bfs_number=int(tc),
                        name=row.get("TerminalName", ""),
                        lindas_uri=LINDAS_MUNICIPALITY_URI.format(bfs=int(tc)),
                    )
                )

        unchanged = (
            len(successors) == 1
            and successors[0].bfs_number == params.bfs_number
            and related[0].get("InitialName") == related[0].get("TerminalName")
        )

        # Mutation path: the change events in range touching this commune's
        # lineage (as origin, or as a target of a merge into a successor).
        mutations = await _fetch_agvch_csv(
            "mutations",
            {
                "includeTerritoryExchange": "false",
                "startPeriod": start,
                "endPeriod": end,
            },
        )
        path = [
            _mutation_step(m)
            for m in mutations
            if m.get("InitialCode") == bfs or (m.get("TerminalCode") or "") in seen
        ][:50]

        if unchanged:
            note = (
                f"BFS-Nummer {bfs} ist zwischen {params.from_date} und "
                f"{params.to_date} unverändert — keine Umschlüsselung nötig."
            )
        else:
            targets = ", ".join(f"{s.name} ({s.bfs_number})" for s in successors)
            note = (
                f"Alte Statistik auf BFS-Nummer {bfs} umschlüsseln auf: {targets}."
            )

        return ResolveHistoricalCommuneResult(
            provenance="live_api",
            bfs_number=params.bfs_number,
            from_date=params.from_date,
            to_date=params.to_date,
            unchanged=unchanged,
            resolves_to=successors,
            mutation_path=path,
            note=note,
        )
    except httpx.HTTPStatusError as e:
        return ResolveHistoricalCommuneResult(
            error=f"AGVCH-API-Fehler {e.response.status_code}",
            hint="Datumsangaben (YYYY-MM-DD) prüfen. Quelle evtl. kurz nicht erreichbar.",
        )
    except Exception:
        _LOGGER.exception("resolve_historical_commune failed")
        return ResolveHistoricalCommuneResult(
            error="Interner Fehler bei der Umschlüsselung.",
            hint="Bitte erneut versuchen.",
        )


@mcp.tool(
    name="list_communes",
    annotations={
        "title": "List communes of a canton",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
@_logged_tool("list_communes")
async def list_communes(params: ListCommunesInput) -> ListCommunesResult:
    """List all communes of a canton, as of a given date.

    Canton membership is derived from the snapshot's Parent chain
    (commune → district → canton), so this reflects the official division
    on `valid_at_date`. Each entry carries its BFS number and LINDAS URI.

    Args:
        params (ListCommunesInput):
            - canton (str): abbreviation ('ZH') or name ('Zürich')
            - valid_at_date (str): ISO date; default today

    Returns:
        ListCommunesResult with the canton's `communes`, sorted by BFS
        number. On error, `error` and `hint` are set.
    """
    try:
        agvch_date = _iso_to_agvch(params.valid_at_date)
        rows, from_cache = await _fetch_snapshot(agvch_date)
        by_hist = _index_by_hist(rows)
        provenance = "cached" if from_cache else "live_api"

        canton_in = params.canton.strip()
        canton_lower = canton_in.lower()
        target: dict[str, str] | None = None
        for r in rows:
            if r.get("Level") != "1":
                continue
            if (
                r.get("ShortName", "").lower() == canton_lower
                or canton_lower in r.get("Name", "").lower()
            ):
                target = r
                break

        if target is None:
            return ListCommunesResult(
                provenance=provenance,
                error=f"Kanton '{canton_in}' nicht gefunden.",
                hint="Kürzel wie 'ZH' oder Name wie 'Zürich' verwenden.",
            )

        target_hist = target["HistoricalCode"]
        communes: list[CommuneEntry] = []
        for r in rows:
            if r.get("Level") != "3":
                continue
            canton_row = _climb_to_canton(r, by_hist)
            if canton_row is not None and canton_row.get("HistoricalCode") == target_hist:
                communes.append(_commune_entry(r, by_hist))
        communes.sort(key=lambda e: e.bfs_number)

        return ListCommunesResult(
            provenance=provenance,
            canton=target.get("Name"),
            canton_abbr=target.get("ShortName"),
            valid_at_date=params.valid_at_date,
            total=len(communes),
            communes=communes,
        )
    except httpx.HTTPStatusError as e:
        return ListCommunesResult(
            error=f"AGVCH-API-Fehler {e.response.status_code}",
            hint="Datum (YYYY-MM-DD) prüfen. Quelle evtl. kurz nicht erreichbar.",
        )
    except Exception:
        _LOGGER.exception("list_communes failed")
        return ListCommunesResult(
            error="Interner Fehler beim Auflisten der Gemeinden.",
            hint="Bitte erneut versuchen.",
        )


# ---------------------------------------------------------------------------
# Reference layer tool — HSSO historical series (Architecture C)
# ---------------------------------------------------------------------------


@mcp.tool(
    name="search_historical_series",
    annotations={
        "title": "Search Historical Statistics of Switzerland (HSSO)",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
@_logged_tool("search_historical_series")
async def search_historical_series(
    params: SearchHistoricalSeriesInput,
) -> SearchHistoricalSeriesResult:
    """Search long-run historical time series (HSSO) by topic.

    Historical Statistics of Switzerland provides long-run series (roughly
    19th–20th century) as static XLSX tables. This tool searches the table
    catalogue by keyword and returns each match with its page and a stable
    XLSX download URL.

    Licence: HSSO is CC BY-NC-SA 3.0 — attribution required, NonCommercial.
    Every response carries that notice in `licence_note`.

    Args:
        params (SearchHistoricalSeriesInput):
            - topic (str): keyword(s); all must match the title
            - period (str): optional period hint (informational only)

    Returns:
        SearchHistoricalSeriesResult with matching `series` (code, title,
        page URL, XLSX URL). On error, `error` and `hint` are set.
    """
    try:
        index, from_cache = await _ensure_hsso_index()
        if not index:
            return SearchHistoricalSeriesResult(
                topic=params.topic,
                period=params.period,
                total_matches=0,
                series=[],
                error="HSSO-Katalog aktuell nicht erreichbar.",
                hint="hsso.ch antwortet nicht. Bitte in einigen Minuten erneut versuchen.",
            )

        terms = params.topic.lower().split()
        matches = [
            e for e in index if all(t in e.title.lower() for t in terms)
        ]
        shown = matches[:25]

        notes: list[str] = []
        if params.period:
            notes.append(
                f"Periodenfilter '{params.period}' ist nur ein Hinweis — HSSO bietet "
                "keine tabellengenaue Periodenfilterung; Periode direkt in der XLSX prüfen."
            )
        if not matches:
            notes.append(
                "Keine Treffer — breitere Stichworte versuchen "
                "(z.B. 'Bevölkerung', 'Preise', 'Verkehr')."
            )
        elif len(matches) > len(shown):
            notes.append(f"{len(matches)} Treffer, angezeigt: {len(shown)}.")

        return SearchHistoricalSeriesResult(
            provenance="cached" if from_cache else "live_api",
            topic=params.topic,
            period=params.period,
            total_matches=len(matches),
            series=shown,
            note=" ".join(notes) if notes else None,
        )
    except Exception:
        _LOGGER.exception("search_historical_series failed")
        return SearchHistoricalSeriesResult(
            error="Interner Fehler bei der HSSO-Suche.",
            hint="hsso.ch evtl. nicht erreichbar. Bitte später erneut versuchen.",
        )


# ---------------------------------------------------------------------------
# Construction & real-estate tools (STAT-TAB theme 09 — Bau- und Wohnungswesen)
# ---------------------------------------------------------------------------
#
# These are consolidated official yearly statistics from STAT-TAB (BFS), the
# same PxWeb source as the rest of the server, so they reuse `_post`,
# `_fetch_metadata_cached` and the shared retry policy. Two findings shape the
# cube choice and the geo resolver (verified live 2026-07-25):
#
# 1. The Gemeinde-level building series was restructured at 2012/2013. The old
#    cubes `_101`–`_104` cover 1995–2012 with a `Kanton (-) / Gemeinde (......)`
#    geo dimension; the current cubes `_105`–`_107` cover 2013–onwards with a
#    `Grossregion (<<) / Kanton (-) / Gemeinde (......)` dimension. Since the
#    default `since_year` is 2015 we query the current cubes.
# 2. PxWeb Gemeinde codes are NOT consistent across cubes. In `_106`/`_107` the
#    value code IS the zero-padded BFS number (`0261`); in `_105` the value code
#    is an opaque sequential id (`160`) and the BFS number appears only inside
#    the label (`......0261 Zürich`). So the resolver matches on the BFS number
#    embedded in the *label*, not on the value code — and each cube is resolved
#    against its own live dimension values, never guessed.

BFS_STATTAB_ATTRIBUTION = (
    "Bundesamt für Statistik (BFS), STAT-TAB — https://www.pxweb.bfs.admin.ch. "
    "Open Government Data, freie Weiterverwendung."
)

# Current (2013–) Gemeinde-level building cubes.
CONSTRUCTION_BUILDINGS_CUBE = "px-x-0904030000_106"  # neu erstellte Gebäude mit Wohnungen
CONSTRUCTION_DWELLINGS_CUBE = "px-x-0904030000_105"  # neu erstellte Wohnungen nach Zimmerzahl
CONSTRUCTION_BUILDINGS_GEO_VAR = "Grossregion (<<) / Kanton (-) / Gemeinde (......)"
CONSTRUCTION_BUILDINGS_TYPE_VAR = "Gebäudetyp"
CONSTRUCTION_ROOMS_VAR = "Anzahl Zimmer"

# Bauinvestitionen und Arbeitsvorrat by region / canton / commune (1994–).
CONSTRUCTION_INVESTMENT_CUBE = "px-x-0904010000_205"
CONSTRUCTION_INVESTMENT_GEO_VAR = "Grossregion (<<) / Kanton (-) / Gemeinde (......)"
CONSTRUCTION_INVESTMENT_WORK_VAR = "Art der Arbeiten"
CONSTRUCTION_INVESTMENT_CATEGORY_VAR = "Kategorie der Bauwerke"
CONSTRUCTION_INVESTMENT_UNIT_VAR = "Beobachtungseinheit"
# Beobachtungseinheit codes: absolute current-year investment + absolute
# next-year Arbeitsvorrat (the monetary leading indicator).
CONSTRUCTION_INVESTMENT_CODE = "kost_j"   # Laufendes Jahr — Absolute Werte
CONSTRUCTION_WORKONHAND_CODE = "arbv_k"   # Folgejahr (Arbeitsvorrat) — Absolute Werte


def _find_var(meta: dict[str, Any], code: str) -> dict[str, Any] | None:
    for v in meta.get("variables", []):
        if v.get("code") == code:
            return v
    return None


def _strip_geo_prefix(text: str) -> str:
    """'......0261 Zürich' → 'Zürich'; '- Waadt' → 'Waadt'; '<< Tessin' → 'Tessin'."""
    return re.sub(r"^[.\-<>\s]*\d*\s*", "", text).strip()


def _iter_jsonstat2(data: dict[str, Any]):
    """Yield (dim_code → value_code dict, value) for every cell, preserving codes.

    Unlike `_format_jsonstat2_as_table` (which yields human labels), this keeps
    the value *codes* so downstream filtering can key on stable identifiers such
    as `Beobachtungseinheit == 'kost_j'` rather than a German label string.
    """
    import itertools

    dims = data.get("id", [])
    dim_info = data.get("dimension", {})
    values = data.get("value", [])
    code_lists: list[list[str]] = []
    for d in dims:
        cats = dim_info.get(d, {}).get("category", {})
        idx = cats.get("index", {})
        if isinstance(idx, dict):
            codes = [c for c, _ in sorted(idx.items(), key=lambda kv: kv[1])]
        else:
            codes = list(idx)
        code_lists.append(codes)
    for combo, val in zip(itertools.product(*code_lists), values):
        yield dict(zip(dims, combo)), val


def _jsonstat2_label(data: dict[str, Any], dim: str, code: str) -> str:
    return (
        data.get("dimension", {})
        .get(dim, {})
        .get("category", {})
        .get("label", {})
        .get(code, code)
    )


def _resolve_municipality_geo(
    meta: dict[str, Any], geo_var: str, bfs: int
) -> tuple[str | None, str | None]:
    """Resolve a BFS commune number to a cube's geo value code + clean name.

    Matches on the BFS number embedded in the commune *label* (`......0261`),
    which is stable across cubes even when the value code is an opaque
    sequential id (see module note). Falls back to a zero-padded code match for
    cubes that expose the BFS number directly as the value code.
    """
    var = _find_var(meta, geo_var)
    if var is None:
        return None, None
    values = var.get("values", [])
    texts = var.get("valueTexts", values)
    for code, text in zip(values, texts):
        m = re.search(r"\.{4,}\s*0*(\d+)", text)  # '......0261 Zürich' → 261
        if m and int(m.group(1)) == bfs:
            return code, _strip_geo_prefix(text)
    for code, text in zip(values, texts):
        cs = code.lstrip(".")
        if cs.isdigit() and len(cs) == 4 and int(cs) == bfs:
            return code, _strip_geo_prefix(text)
    return None, None


def _resolve_investment_geo(
    meta: dict[str, Any], level: str, code: str
) -> tuple[str | None, str | None]:
    """Resolve a (level, code) pair to the investment cube's geo value code.

    - gemeinde: `code` is a BFS number → matched via the label-embedded number.
    - kanton:   `code` is a two-letter abbreviation (e.g. 'ZH').
    - grossregion: `code` is a region id ('R1'…'R7') or a region name.
    All fall back to a case-insensitive name substring match.
    """
    var = _find_var(meta, CONSTRUCTION_INVESTMENT_GEO_VAR)
    if var is None:
        return None, None
    values = var.get("values", [])
    texts = var.get("valueTexts", values)
    wanted = code.strip()

    if level == "gemeinde":
        if wanted.isdigit():
            bfs = int(wanted)
            for c, t in zip(values, texts):
                m = re.search(r"\.{4,}\s*0*(\d+)", t)
                if m and int(m.group(1)) == bfs:
                    return c, _strip_geo_prefix(t)
            for c, t in zip(values, texts):
                cs = c.lstrip(".")
                if cs.isdigit() and len(cs) == 4 and int(cs) == bfs:
                    return c, _strip_geo_prefix(t)
        return None, None

    # kanton / grossregion: exact code match first, then name substring.
    for c, t in zip(values, texts):
        if c.upper() == wanted.upper():
            return c, _strip_geo_prefix(t)
    wl = wanted.lower()
    for c, t in zip(values, texts):
        if wl in _strip_geo_prefix(t).lower():
            return c, _strip_geo_prefix(t)
    return None, None


class ConstructionActivityInput(BaseModel):
    model_config = ConfigDict(extra="forbid")
    municipality_bfs: int = Field(
        ...,
        description=(
            "BFS commune number (e.g. 261 for Zürich). Resolve names to numbers "
            "with lookup_commune first if needed."
        ),
        ge=1,
        le=9999,
    )
    since_year: int = Field(
        default=2015,
        description=(
            "Earliest year to include (inclusive). The current Gemeinde-level "
            "building series starts in 2013; older years live in the discontinued "
            "cubes px-x-0904030000_101/_104 (1995–2012) and are not queried here."
        ),
        ge=2013,
        le=2100,
    )


class ConstructionInvestmentInput(BaseModel):
    model_config = ConfigDict(extra="forbid")
    level: Literal["grossregion", "kanton", "gemeinde"] = Field(
        ...,
        description=(
            "Geographic level of `code`: 'grossregion' (R1–R7), 'kanton' "
            "(abbreviation like 'ZH'), or 'gemeinde' (BFS commune number)."
        ),
    )
    code: str = Field(
        ...,
        description=(
            "Region/canton/commune code matching `level`: e.g. 'R1' or "
            "'Genferseeregion' (grossregion), 'ZH' (kanton), '261' (gemeinde)."
        ),
        min_length=1,
        max_length=40,
    )
    since_year: int = Field(
        default=2015,
        description="Earliest year to include (inclusive). Series starts in 1994.",
        ge=1994,
        le=2100,
    )


class ConstructionActivityYear(BaseModel):
    year: int
    new_buildings: int | None = None  # neu erstellte Gebäude mit Wohnungen
    new_dwellings: int | None = None  # neu erstellte Wohnungen (Total)
    dwellings_by_rooms: dict[str, int | None] | None = None  # room label → count


class ConstructionActivityResult(BaseModel):
    source: str = BFS_STATTAB_ATTRIBUTION
    provenance: str | None = None
    error: str | None = None
    hint: str | None = None
    municipality_bfs: int | None = None
    municipality_name: str | None = None
    since_year: int | None = None
    table_ids: list[str] | None = None
    years: list[ConstructionActivityYear] | None = None
    cross_validation: str | None = None
    note: str | None = None


class ConstructionInvestmentYear(BaseModel):
    year: int
    investment: float | None = None    # Bauinvestitionen, laufendes Jahr (absolut)
    work_on_hand: float | None = None  # Arbeitsvorrat Folgejahr (absolut)


class ConstructionInvestmentResult(BaseModel):
    source: str = BFS_STATTAB_ATTRIBUTION
    provenance: str | None = None
    error: str | None = None
    hint: str | None = None
    level: str | None = None
    code: str | None = None
    region_name: str | None = None
    since_year: int | None = None
    table_id: str | None = None
    unit: str | None = None
    years: list[ConstructionInvestmentYear] | None = None
    note: str | None = None


@mcp.tool(
    name="bfs_construction_activity",
    annotations={
        "title": "Swiss Construction Activity (new buildings & dwellings)",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
@_logged_tool("bfs_construction_activity")
async def bfs_construction_activity(
    params: ConstructionActivityInput,
) -> ConstructionActivityResult:
    """Yearly new buildings and new dwellings for a commune, with room-size mix.

    Returns the consolidated official annual construction statistics (BFS
    STAT-TAB theme 09) for one commune: newly built buildings with dwellings
    (px-x-0904030000_106) and newly built dwellings broken down by number of
    rooms (px-x-0904030000_105), as a per-year series from `since_year`.

    Note: this is the *consolidated official yearly* statistic. For up-to-date
    building-register states and the construction pipeline (Baugesuche /
    Bauvorhaben), see the `swiss-housing-mcp` server — the overlap is deliberate
    so the two sources can be cross-validated.

    Args:
        params (ConstructionActivityInput):
            - municipality_bfs (int): BFS commune number, e.g. 261 (Zürich)
            - since_year (int): earliest year, inclusive (default 2015)

    Returns:
        ConstructionActivityResult with a `years` series (new_buildings,
        new_dwellings, dwellings_by_rooms). On error, `error`/`hint` are set.
    """
    try:
        b_meta = await _fetch_metadata_cached(CONSTRUCTION_BUILDINGS_CUBE, "de")
        geo_code, name = _resolve_municipality_geo(
            b_meta, CONSTRUCTION_BUILDINGS_GEO_VAR, params.municipality_bfs
        )
        if geo_code is None:
            return ConstructionActivityResult(
                municipality_bfs=params.municipality_bfs,
                error=f"Gemeinde mit BFS-Nummer {params.municipality_bfs} nicht in der Baustatistik gefunden.",
                hint=(
                    "BFS-Nummer mit lookup_commune prüfen. Die aktuelle "
                    "Gemeinde-Baustatistik deckt Gemeinden ab 2013 ab."
                ),
            )

        # New buildings with dwellings (Gebäudetyp total).
        buildings_body = {
            "query": [
                {
                    "code": CONSTRUCTION_BUILDINGS_GEO_VAR,
                    "selection": {"filter": "item", "values": [geo_code]},
                },
                {
                    "code": CONSTRUCTION_BUILDINGS_TYPE_VAR,
                    "selection": {"filter": "item", "values": ["0"]},
                },
            ],
            "response": {"format": "json-stat2"},
        }
        b_data = await _post(
            _build_data_url(CONSTRUCTION_BUILDINGS_CUBE, "de"), buildings_body
        )
        buildings_by_year: dict[str, int | None] = {}
        for dims, val in _iter_jsonstat2(b_data):
            year = _jsonstat2_label(b_data, "Jahr", dims.get("Jahr", ""))
            buildings_by_year[year] = val

        # New dwellings by number of rooms — resolved against this cube's own
        # (differently coded) geo dimension.
        d_meta = await _fetch_metadata_cached(CONSTRUCTION_DWELLINGS_CUBE, "de")
        d_geo_code, _ = _resolve_municipality_geo(
            d_meta, CONSTRUCTION_BUILDINGS_GEO_VAR, params.municipality_bfs
        )
        dwellings_total: dict[str, int | None] = {}
        rooms_by_year: dict[str, dict[str, int | None]] = {}
        if d_geo_code is not None:
            rooms_var = _find_var(d_meta, CONSTRUCTION_ROOMS_VAR)
            room_codes = rooms_var.get("values", []) if rooms_var else []
            dwellings_body = {
                "query": [
                    {
                        "code": CONSTRUCTION_BUILDINGS_GEO_VAR,
                        "selection": {"filter": "item", "values": [d_geo_code]},
                    },
                    {
                        "code": CONSTRUCTION_ROOMS_VAR,
                        "selection": {"filter": "item", "values": room_codes},
                    },
                ],
                "response": {"format": "json-stat2"},
            }
            d_data = await _post(
                _build_data_url(CONSTRUCTION_DWELLINGS_CUBE, "de"), dwellings_body
            )
            for dims, val in _iter_jsonstat2(d_data):
                year = _jsonstat2_label(d_data, "Jahr", dims.get("Jahr", ""))
                room_code = dims.get(CONSTRUCTION_ROOMS_VAR, "")
                if room_code == "0":  # Wohnungen — Total
                    dwellings_total[year] = val
                else:
                    label = _jsonstat2_label(d_data, CONSTRUCTION_ROOMS_VAR, room_code)
                    rooms_by_year.setdefault(year, {})[label] = val

        all_years = sorted(
            {y for y in buildings_by_year} | {y for y in dwellings_total},
            key=lambda y: int(y) if y.isdigit() else 0,
        )
        years: list[ConstructionActivityYear] = []
        for y in all_years:
            if not y.isdigit() or int(y) < params.since_year:
                continue
            years.append(
                ConstructionActivityYear(
                    year=int(y),
                    new_buildings=buildings_by_year.get(y),
                    new_dwellings=dwellings_total.get(y),
                    dwellings_by_rooms=rooms_by_year.get(y) or None,
                )
            )

        note = None
        if d_geo_code is None:
            note = (
                "Zimmerzahl-Aufschlüsselung für diese Gemeinde nicht verfügbar; "
                "nur Gebäudezahlen zurückgegeben."
            )
        return ConstructionActivityResult(
            provenance="live_api",
            municipality_bfs=params.municipality_bfs,
            municipality_name=name,
            since_year=params.since_year,
            table_ids=[CONSTRUCTION_BUILDINGS_CUBE, CONSTRUCTION_DWELLINGS_CUBE],
            years=years,
            cross_validation=(
                "Konsolidierte amtliche Jahresstatistik. Für tagesaktuelle "
                "Registerstände und die Bau-Pipeline: swiss-housing-mcp "
                "(bewusste Redundanz zur Cross-Validation)."
            ),
            note=note,
        )
    except httpx.HTTPStatusError as e:
        if e.response.status_code == 400:
            return ConstructionActivityResult(
                error="Ungültige Abfrage (HTTP 400).",
                hint="BFS-Nummer prüfen; STAT-TAB hat die Auswahl abgelehnt.",
            )
        return ConstructionActivityResult(error=f"API-Fehler {e.response.status_code}")
    except Exception:
        _LOGGER.exception("bfs_construction_activity failed")
        return ConstructionActivityResult(
            error="Interner Fehler beim Abruf der Bautätigkeit.",
            hint="Bitte erneut versuchen.",
        )


@mcp.tool(
    name="bfs_construction_investment",
    annotations={
        "title": "Swiss Construction Investment & Arbeitsvorrat",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
@_logged_tool("bfs_construction_investment")
async def bfs_construction_investment(
    params: ConstructionInvestmentInput,
) -> ConstructionInvestmentResult:
    """Yearly building investment and Arbeitsvorrat for a region/canton/commune.

    Returns building investment (Bauinvestitionen, current year) alongside the
    Arbeitsvorrat (work on hand for the following year) from BFS STAT-TAB
    px-x-0904010000_205, as a per-year series from `since_year`. The
    Arbeitsvorrat is the monetary leading indicator: it signals next year's
    construction volume before it is realised.

    Args:
        params (ConstructionInvestmentInput):
            - level (str): 'grossregion', 'kanton', or 'gemeinde'
            - code (str): region/canton/commune code matching `level`
            - since_year (int): earliest year, inclusive (default 2015)

    Returns:
        ConstructionInvestmentResult with a `years` series (investment,
        work_on_hand) in 1000 CHF. On error, `error`/`hint` are set.
    """
    try:
        meta = await _fetch_metadata_cached(CONSTRUCTION_INVESTMENT_CUBE, "de")
        geo_code, region_name = _resolve_investment_geo(meta, params.level, params.code)
        if geo_code is None:
            return ConstructionInvestmentResult(
                level=params.level,
                code=params.code,
                error=f"Kein Eintrag für {params.level}='{params.code}' gefunden.",
                hint=(
                    "grossregion: 'R1'–'R7'; kanton: Kürzel wie 'ZH'; "
                    "gemeinde: BFS-Nummer (lookup_commune)."
                ),
            )

        body = {
            "query": [
                {
                    "code": CONSTRUCTION_INVESTMENT_GEO_VAR,
                    "selection": {"filter": "item", "values": [geo_code]},
                },
                {
                    "code": CONSTRUCTION_INVESTMENT_WORK_VAR,
                    "selection": {"filter": "item", "values": ["0"]},  # Total
                },
                {
                    "code": CONSTRUCTION_INVESTMENT_CATEGORY_VAR,
                    "selection": {"filter": "item", "values": ["0"]},  # Total
                },
                {
                    "code": CONSTRUCTION_INVESTMENT_UNIT_VAR,
                    "selection": {
                        "filter": "item",
                        "values": [
                            CONSTRUCTION_INVESTMENT_CODE,
                            CONSTRUCTION_WORKONHAND_CODE,
                        ],
                    },
                },
            ],
            "response": {"format": "json-stat2"},
        }
        data = await _post(_build_data_url(CONSTRUCTION_INVESTMENT_CUBE, "de"), body)

        investment: dict[str, float | None] = {}
        work_on_hand: dict[str, float | None] = {}
        for dims, val in _iter_jsonstat2(data):
            year = _jsonstat2_label(data, "Jahr", dims.get("Jahr", ""))
            unit = dims.get(CONSTRUCTION_INVESTMENT_UNIT_VAR, "")
            if unit == CONSTRUCTION_INVESTMENT_CODE:
                investment[year] = val
            elif unit == CONSTRUCTION_WORKONHAND_CODE:
                work_on_hand[year] = val

        all_years = sorted(
            {y for y in investment} | {y for y in work_on_hand},
            key=lambda y: int(y) if y.isdigit() else 0,
        )
        years = [
            ConstructionInvestmentYear(
                year=int(y),
                investment=investment.get(y),
                work_on_hand=work_on_hand.get(y),
            )
            for y in all_years
            if y.isdigit() and int(y) >= params.since_year
        ]

        return ConstructionInvestmentResult(
            provenance="live_api",
            level=params.level,
            code=params.code,
            region_name=region_name,
            since_year=params.since_year,
            table_id=CONSTRUCTION_INVESTMENT_CUBE,
            unit="1000 CHF",
            years=years,
            note=(
                "Werte in 1000 CHF. Arbeitsvorrat = Bauvolumen des Folgejahres "
                "(monetärer Frühindikator)."
            ),
        )
    except httpx.HTTPStatusError as e:
        if e.response.status_code == 400:
            return ConstructionInvestmentResult(
                error="Ungültige Abfrage (HTTP 400).",
                hint="level/code-Kombination prüfen; STAT-TAB hat die Auswahl abgelehnt.",
            )
        return ConstructionInvestmentResult(error=f"API-Fehler {e.response.status_code}")
    except Exception:
        _LOGGER.exception("bfs_construction_investment failed")
        return ConstructionInvestmentResult(
            error="Interner Fehler beim Abruf der Bauinvestitionen.",
            hint="Bitte erneut versuchen.",
        )


# ---------------------------------------------------------------------------
# Price indices — IMPI & Baupreisindex (BFS DAM asset API + opendata.swiss CKAN)
# ---------------------------------------------------------------------------
#
# The residential property price index (IMPI) and the construction price index
# (Baupreisindex) are NOT in STAT-TAB. They are published as BFS DAM assets
# (dam-api.bfs.admin.ch) whose dataset metadata lives on opendata.swiss (CKAN).
# Findings that shape this tool (verified live 2026-07-25):
#
# 1. CKAN 403-without-User-Agent: ckan.opendata.swiss rejects requests with a
#    default httpx/curl User-Agent (HTTP 403). A custom UA is mandatory, so
#    every CKAN/DAM call here goes through `_get_json_ua` / `_get_bytes_ua`,
#    which set `CKAN_USER_AGENT`.
# 2. DAM assets mix formats. IMPI ships only PDF + HTML — there is NO
#    machine-readable series — so for `index="impi"` this tool returns the
#    official source links plus an explicit limitation, not parsed values.
#    The Baupreisindex ships an XLSX; the tool selects it by content-type
#    (skipping PDFs) and parses the national semi-annual series with openpyxl.

CKAN_API_BASE = "https://ckan.opendata.swiss/api/3/action"

try:  # UA carries the real package version so BFS can attribute traffic.
    from importlib.metadata import version as _pkg_version

    _UA_VERSION = _pkg_version("swiss-statistics-mcp")
except Exception:  # pragma: no cover - fallback when metadata is unavailable
    _UA_VERSION = "0.0.0"
CKAN_USER_AGENT = f"swiss-statistics-mcp/{_UA_VERSION}"

BFS_PRICE_ATTRIBUTION = (
    "Bundesamt für Statistik (BFS) — Preisindizes via opendata.swiss (CKAN) "
    "und BFS DAM-Asset-API. Open Government Data, freie Weiterverwendung."
)

PRICE_INDEX_CACHE_TTL = int(os.environ.get("MCP_PRICE_INDEX_TTL", "86400"))  # 24h
_XLSX_CONTENT_TYPES = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
)

# CKAN full-text queries per index. The dataset/resource is then selected from
# the live metadata (never a hard-coded DAM asset id, which changes on
# republish).
_PRICE_INDEX_QUERY = {
    "baupreisindex": "Baupreisindex Multibasen Grossregion Objekttyp",
    "impi": "Wohnimmobilienpreisindex",
}
# Month names (DE) → 2-digit month, for the semi-annual Baupreisindex header.
_BPI_MONTHS = {
    "Januar": "01",
    "April": "04",
    "Juli": "07",
    "Oktober": "10",
}

# Cache the fully parsed result per index; `since_year` filters at return time.
_price_index_cache: dict[str, tuple[float, PriceIndexResult]] = {}


async def _get_json_ua(url: str) -> Any:
    """GET JSON with the mandatory custom User-Agent (CKAN 403 guard)."""

    async def _do() -> Any:
        async with httpx.AsyncClient(timeout=HTTP_TIMEOUT, follow_redirects=True) as client:
            resp = await client.get(url, headers={"User-Agent": CKAN_USER_AGENT})
            resp.raise_for_status()
            return resp.json()

    return await _retrying_http(_do)


async def _get_bytes_ua(url: str) -> tuple[str, bytes]:
    """GET raw bytes + content-type with the custom User-Agent, following redirects."""

    async def _do() -> tuple[str, bytes]:
        async with httpx.AsyncClient(timeout=HTTP_TIMEOUT, follow_redirects=True) as client:
            resp = await client.get(url, headers={"User-Agent": CKAN_USER_AGENT})
            resp.raise_for_status()
            return resp.headers.get("content-type", ""), resp.content

    return await _retrying_http(_do)


def _ckan_localized(value: Any) -> str:
    """CKAN titles are `{de,fr,it,en}` dicts (or plain strings) — prefer German."""
    if isinstance(value, dict):
        for lang in ("de", "fr", "it", "en"):
            if value.get(lang):
                return value[lang]
        return next((v for v in value.values() if v), "")
    return value or ""


def _parse_baupreisindex_xlsx(
    content: bytes,
) -> tuple[str, str, str, list[tuple[str, float]]] | None:
    """Parse the national Baugewerbe-Total series from a Baupreisindex XLSX.

    The workbook has one sheet per index base (named by year); the latest is
    used. Region/object rows are keyed by stable codes (`<REG_01>` = national,
    `<OBJ_02>` = Baugewerbe Total) rather than language-dependent labels.

    Returns (base_sheet, base_label, object_label, [(period, value)]) or None if
    the expected structure isn't present (so the caller can fall back cleanly).
    """
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    base_sheets = sorted((s for s in wb.sheetnames if s.isdigit()), key=int)
    if not base_sheets:
        return None
    base = base_sheets[-1]
    rows = list(wb[base].iter_rows(values_only=True))

    base_label = base
    for r in rows[:10]:
        if r and isinstance(r[0], str) and r[0].startswith("<BASE_"):
            base_label = r[1] if (len(r) > 1 and r[1]) else base
            break

    # Locate the month-header row (contains a known month name) + the year row.
    month_row_idx = None
    for i, r in enumerate(rows[:15]):
        if any(isinstance(c, str) and c in _BPI_MONTHS for c in r):
            month_row_idx = i
            break
    if month_row_idx is None or month_row_idx + 1 >= len(rows):
        return None
    months = rows[month_row_idx]
    years = rows[month_row_idx + 1]
    periods: dict[int, str] = {}
    for j, (mn, yr) in enumerate(zip(months, years)):
        if isinstance(mn, str) and mn in _BPI_MONTHS and yr not in (None, ""):
            try:
                periods[j] = f"{int(yr)}-{_BPI_MONTHS[mn]}"
            except (TypeError, ValueError):
                continue
    if not periods:
        return None

    cur_region: str | None = None
    for r in rows:
        code = r[0]
        if isinstance(code, str) and code.startswith("<REG_"):
            cur_region = code
        if cur_region == "<REG_01>" and code == "<OBJ_02>":
            obj_label = (r[1] or "Baugewerbe: Total") if len(r) > 1 else "Baugewerbe: Total"
            series = [
                (periods[j], round(float(r[j]), 2))
                for j in sorted(periods)
                if j < len(r) and isinstance(r[j], (int, float))
            ]
            if series:
                return base, base_label, str(obj_label).replace("\xa0", " ").strip(), series
            return None
    return None


class PriceIndexInput(BaseModel):
    model_config = ConfigDict(extra="forbid")
    index: Literal["impi", "baupreisindex"] = Field(
        ...,
        description=(
            "'baupreisindex' (construction price index — parsed national series) "
            "or 'impi' (residential property price index — BFS publishes it only "
            "as PDF, so this returns the official source links, not values)."
        ),
    )
    since_year: int | None = Field(
        default=None,
        description="Optional earliest year (inclusive) to include in the series.",
        ge=1900,
        le=2100,
    )


class PriceIndexPoint(BaseModel):
    period: str  # 'YYYY-MM' (semi-annual for the Baupreisindex)
    value: float


class PriceIndexResult(BaseModel):
    source: str = BFS_PRICE_ATTRIBUTION
    provenance: str | None = None
    error: str | None = None
    hint: str | None = None
    index: str | None = None
    title: str | None = None
    dataset: str | None = None
    base: str | None = None  # e.g. 'Basis Oktober 2025 = 100'
    coverage: str | None = None  # e.g. 'Schweiz — Baugewerbe: Total'
    series: list[PriceIndexPoint] | None = None
    source_links: list[str] | None = None  # for IMPI / discovery
    note: str | None = None


async def _load_price_index(index: str) -> PriceIndexResult:
    """Fetch + parse a price index (uncached); results are cached by the tool."""
    query = _PRICE_INDEX_QUERY[index]
    search = await _get_json_ua(
        f"{CKAN_API_BASE}/package_search?q={quote_plus(query)}&rows=10"
    )
    datasets = search.get("result", {}).get("results", [])
    if not datasets:
        return PriceIndexResult(
            index=index,
            error="Kein passendes opendata.swiss-Dataset gefunden.",
            hint="CKAN evtl. kurz nicht erreichbar. Bitte später erneut versuchen.",
        )

    if index == "impi":
        # IMPI: PDF/HTML only — return official source links, no parsed series.
        ds = next(
            (d for d in datasets if d.get("name") == "schweizerischer-wohnimmobilienpreisindex-impi"),
            datasets[0],
        )
        links: list[str] = []
        for res in ds.get("resources", []):
            fmt = (res.get("format") or "").upper()
            url = res.get("url") or res.get("download_url")
            if url and fmt in ("PDF", "HTML"):
                links.append(url)
        return PriceIndexResult(
            provenance="live_api",
            index=index,
            title=_ckan_localized(ds.get("title")),
            dataset=ds.get("name"),
            source_links=links[:8],
            note=(
                "Der Wohnimmobilienpreisindex (IMPI) wird vom BFS nur als PDF/HTML "
                "publiziert — keine maschinenlesbare Zeitreihe verfügbar. Die "
                "Indexwerte den verlinkten Quellen (PDF) entnehmen."
            ),
        )

    # Baupreisindex: find an XLSX resource, verify content-type, parse.
    candidates: list[str] = []
    for ds in datasets:
        for res in ds.get("resources", []):
            fmt = (res.get("format") or "").upper()
            url = res.get("url") or res.get("download_url")
            if url and (fmt in ("XLSX", "XLS") or "dam-api.bfs.admin.ch" in url):
                candidates.append(url)

    for url in candidates[:6]:
        try:
            content_type, content = await _get_bytes_ua(url)
        except Exception:
            _LOGGER.warning("price_index asset fetch failed for %s", url, exc_info=True)
            continue
        if not any(ct in content_type for ct in _XLSX_CONTENT_TYPES):
            continue  # skip PDFs and other non-spreadsheet assets
        parsed = _parse_baupreisindex_xlsx(content)
        if parsed is None:
            continue
        base, base_label, obj_label, series = parsed
        ds = datasets[0]
        return PriceIndexResult(
            provenance="live_api",
            index=index,
            title=_ckan_localized(ds.get("title")),
            dataset=ds.get("name"),
            base=base_label,
            coverage=f"Schweiz — {obj_label}",
            series=[PriceIndexPoint(period=p, value=v) for p, v in series],
            note=(
                "Halbjährliche Indexreihe (April/Oktober) für die Schweiz gesamt "
                "(Baugewerbe: Total). Regionale und objektspezifische Reihen "
                "stehen in der Quell-XLSX zur Verfügung."
            ),
        )

    return PriceIndexResult(
        index=index,
        error="Keine parsbare Baupreisindex-XLSX gefunden.",
        hint=(
            "Die DAM-Assets liefern evtl. nur PDF, oder die XLSX-Struktur hat sich "
            "geändert. Quelle manuell auf opendata.swiss prüfen."
        ),
    )


@mcp.tool(
    name="bfs_price_index",
    annotations={
        "title": "Swiss Price Indices (IMPI / Baupreisindex)",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
@_logged_tool("bfs_price_index")
async def bfs_price_index(params: PriceIndexInput) -> PriceIndexResult:
    """Swiss price indices not carried by STAT-TAB: Baupreisindex & IMPI.

    - `baupreisindex`: the construction price index — returns the national
      semi-annual index series (Schweiz, Baugewerbe Total), parsed from the BFS
      DAM XLSX selected via opendata.swiss (CKAN) metadata.
    - `impi`: the residential property price index — BFS publishes this only as
      PDF/HTML, so this returns the official source links plus an explicit
      limitation rather than parsed values.

    Data flows through opendata.swiss (CKAN), which rejects default User-Agents
    with HTTP 403; a custom User-Agent is always sent. Results are cached for
    24h.

    Args:
        params (PriceIndexInput):
            - index (str): 'baupreisindex' or 'impi'
            - since_year (int | None): optional earliest year to include

    Returns:
        PriceIndexResult with `series` (baupreisindex) or `source_links` (impi).
        On error, `error`/`hint` are set.
    """
    try:
        now = time.time()
        cached = _price_index_cache.get(params.index)
        if cached is not None and (now - cached[0]) < PRICE_INDEX_CACHE_TTL:
            result = cached[1].model_copy()
            result.provenance = "cached"
        else:
            result = await _load_price_index(params.index)
            if result.error is None:
                _price_index_cache[params.index] = (now, result)

        if params.since_year is not None and result.series:
            filtered = [
                p for p in result.series if int(p.period[:4]) >= params.since_year
            ]
            result = result.model_copy(update={"series": filtered})
        return result
    except httpx.HTTPStatusError as e:
        return PriceIndexResult(
            index=params.index,
            error=f"API-Fehler {e.response.status_code}",
            hint="opendata.swiss / BFS DAM evtl. kurz nicht erreichbar.",
        )
    except Exception:
        _LOGGER.exception("bfs_price_index failed")
        return PriceIndexResult(
            index=params.index,
            error="Interner Fehler beim Abruf des Preisindex.",
            hint="Bitte erneut versuchen.",
        )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import os
    import sys

    if "--http" in sys.argv:
        port_idx = sys.argv.index("--port") + 1 if "--port" in sys.argv else None
        port     = int(sys.argv[port_idx]) if port_idx else 8000

        # Default to loopback. The server has no authentication; exposing it on
        # 0.0.0.0 turns it into an open proxy to the BFS API. Set MCP_HOST or
        # pass --host explicitly to bind elsewhere (e.g. behind a reverse proxy
        # with access control).
        host_idx = sys.argv.index("--host") + 1 if "--host" in sys.argv else None
        host     = sys.argv[host_idx] if host_idx else os.environ.get("MCP_HOST", "127.0.0.1")

        mcp.settings.host = host
        mcp.settings.port = port
        mcp.run(transport="streamable-http")
    else:
        mcp.run()
